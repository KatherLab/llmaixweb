import csv
import datetime
import io
import json
import re
import zipfile
from typing import Any, List, cast

import pandas as pd
from fastapi import (
    APIRouter,
    Body,
    Depends,
    File,
    Form,
    HTTPException,
    Query,
    UploadFile,
)
from fastapi.responses import Response, StreamingResponse
from pydantic import ValidationError
from sqlalchemy import and_, delete, distinct, func, or_, select
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.orm import Session, selectinload
from starlette import status
from thefuzz import fuzz

from .... import models, schemas
from ....core.config import settings
from ....core.security import get_admin_user, get_current_user
from ....dependencies import (
    calculate_file_hash,
    get_db,
    get_file,
    remove_file,
    save_file,
)
from ....models.project import document_set_association
from ....utils.enums import FileCreator, FileType
from ....utils.helpers import (
    extract_field_types_from_schema,
    flatten_dict,
    validate_prompt,
)
from ....utils.info_extraction import (
    get_available_models,
    test_api_connection,
    test_llm_connection,
    test_model_with_schema,
)

router = APIRouter()


def check_project_access(
    project_id: int, current_user: models.User, db: Session, permission: str = "read"
) -> models.Project:
    """Check if user has access to project."""
    project = db.execute(
        select(models.Project).where(models.Project.id == project_id)
    ).scalar_one_or_none()

    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    # Admin has full access
    if current_user.role == "admin":
        return project

    # Owner has full access
    if project.owner_id == current_user.id:
        return project

    # For non-owners, check specific permissions if needed
    # You could extend this with a project_members table for shared projects
    raise HTTPException(
        status_code=403, detail=f"Not authorized to {permission} this project"
    )


@router.get("/", response_model=list[schemas.Project])
def get_projects(
    *,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
) -> list[schemas.Project]:
    if current_user.role == "admin":
        projects = list(db.execute(select(models.Project)).scalars().all())
    else:
        projects = list(
            db.execute(
                select(models.Project).where(models.Project.owner_id == current_user.id)
            )
            .scalars()
            .all()
        )
    return projects


@router.get("/{project_id}", response_model=schemas.Project)
def get_project(
    *,
    db: Session = Depends(get_db),
    project_id: int,
    current_user: models.User = Depends(get_current_user),
) -> schemas.Project:
    project: models.Project | None = db.execute(
        select(models.Project).where(models.Project.id == project_id)
    ).scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    if current_user.role != "admin" and project.owner_id != current_user.id:
        raise HTTPException(
            status_code=403, detail="Not authorized to access this project"
        )

    return project


@router.post("/", response_model=schemas.Project)
def create_project(
    *,
    db: Session = Depends(get_db),
    project: schemas.ProjectCreate,
    current_user: models.User = Depends(get_current_user),
) -> schemas.Project:
    if not current_user.role == "user":
        if not project.owner_id:
            project.owner_id = current_user.id
        elif project.owner_id != current_user.id:
            raise HTTPException(
                status_code=403,
                detail="Not authorized to create project for another user",
            )
    else:
        if not project.owner_id:
            project.owner_id = current_user.id

    new_project = models.Project(**project.model_dump())

    db.add(new_project)
    db.commit()
    db.refresh(new_project)

    return schemas.Project.model_validate(new_project)


@router.put("/{project_id}", response_model=schemas.Project)
def update_project(
    *,
    db: Session = Depends(get_db),
    project_id: int,
    project: schemas.ProjectUpdate,
    current_user: models.User = Depends(get_current_user),
) -> schemas.Project:
    existing_project: models.Project | None = db.execute(
        select(models.Project).where(models.Project.id == project_id)
    ).scalar_one_or_none()

    if not existing_project:
        raise HTTPException(status_code=404, detail="Project not found")

    if current_user.role != "admin" and existing_project.owner_id != current_user.id:
        raise HTTPException(
            status_code=403, detail="Not authorized to update this project"
        )

    for key, value in project.model_dump(exclude_unset=True).items():
        setattr(existing_project, key, value)

    db.add(existing_project)
    db.commit()
    db.refresh(existing_project)

    return schemas.Project.model_validate(existing_project)


@router.delete("/{project_id}", response_model=schemas.Project)
def delete_project(
    *,
    db: Session = Depends(get_db),
    project_id: int,
    current_user: models.User = Depends(get_current_user),
) -> schemas.Project:
    existing_project: models.Project | None = db.execute(
        select(models.Project).where(models.Project.id == project_id)
    ).scalar_one_or_none()

    if not existing_project:
        raise HTTPException(status_code=404, detail="Project not found")

    if current_user.role != "admin" and existing_project.owner_id != current_user.id:
        raise HTTPException(
            status_code=403, detail="Not authorized to delete this project"
        )

    # TODO: Make sure to handle any related resources (e.g., files, documents) before deleting the project

    db.delete(existing_project)
    db.commit()

    return schemas.Project.model_validate(existing_project)


@router.get("/{project_id}/file", response_model=list[schemas.File])
def get_project_files(
    *,
    db: Session = Depends(get_db),
    project_id: int,
    # Filter parameters
    search: str | None = Query(None, description="Search in filename"),
    file_type: str | None = Query(None, description="Filter by file type"),
    file_creator: FileCreator | None = Query(None),
    date_from: datetime.datetime | None = Query(None),
    date_to: datetime.datetime | None = Query(None),
    min_size: int | None = Query(None, description="Minimum file size in bytes"),
    max_size: int | None = Query(None, description="Maximum file size in bytes"),
    # Pagination
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=1000),
    current_user: models.User = Depends(get_current_user),
) -> list[schemas.File]:
    """Get project files with advanced filtering"""
    check_project_access(project_id, current_user, db)

    query = select(models.File).where(models.File.project_id == project_id)

    # Apply filters
    if search:
        query = query.where(models.File.file_name.ilike(f"%{search}%"))
    if file_type:
        query = query.where(models.File.file_type == file_type)
    if file_creator is not None:
        query = query.where(models.File.file_creator == file_creator)
    if date_from:
        query = query.where(models.File.created_at >= date_from)
    if date_to:
        query = query.where(models.File.created_at <= date_to)
    if min_size is not None:
        query = query.where(models.File.file_size >= min_size)
    if max_size is not None:
        query = query.where(models.File.file_size <= max_size)

    # Order by created_at desc and apply pagination
    query = query.order_by(models.File.created_at.desc()).offset(skip).limit(limit)

    files = list(db.execute(query).scalars().all())
    return [schemas.File.model_validate(file) for file in files]


@router.get("/{project_id}/file/stats", response_model=dict)
def get_file_stats(
    *,
    db: Session = Depends(get_db),
    project_id: int,
    file_creator: FileCreator | None = Query(None),
    current_user: models.User = Depends(get_current_user),
) -> dict:
    """Get file statistics for the project"""
    check_project_access(project_id, current_user, db)

    # Base query
    base_query = select(models.File).where(models.File.project_id == project_id)
    if file_creator is not None:
        base_query = base_query.where(models.File.file_creator == file_creator)

    # Get total count and size
    stats = db.execute(
        select(
            func.count(models.File.id).label("total_files"),
            func.sum(models.File.file_size).label("total_size"),
            func.count(distinct(models.File.file_hash)).label("unique_files"),
        ).where(
            models.File.project_id == project_id,
            models.File.file_creator == file_creator if file_creator else True,
        )
    ).first()

    # Get files by type
    type_query = select(
        models.File.file_type,
        func.count(models.File.id).label("count"),
        func.sum(models.File.file_size).label("size"),
    ).where(models.File.project_id == project_id)

    if file_creator is not None:
        type_query = type_query.where(models.File.file_creator == file_creator)

    type_stats = db.execute(type_query.group_by(models.File.file_type)).all()

    # Get recent files (last 7 days)
    week_ago = datetime.datetime.now(datetime.UTC) - datetime.timedelta(days=7)
    recent_query = select(func.count(models.File.id)).where(
        and_(models.File.project_id == project_id, models.File.created_at >= week_ago)
    )

    if file_creator is not None:
        recent_query = recent_query.where(models.File.file_creator == file_creator)

    recent_count = db.execute(recent_query).scalar()

    return {
        "total_files": stats.total_files or 0,
        "total_size": stats.total_size or 0,
        "unique_files": stats.unique_files or 0,
        "recent_files": recent_count or 0,
        "duplicates": (stats.total_files or 0) - (stats.unique_files or 0),
        "by_type": [
            {"type": t.file_type, "count": t.count, "size": t.size or 0}
            for t in type_stats
        ],
    }


@router.get("/{project_id}/file/{file_id}", response_model=schemas.File)
def get_project_file(
    *,
    db: Session = Depends(get_db),
    project_id: int,
    file_id: int,
    current_user: models.User = Depends(get_current_user),
) -> schemas.File:
    project: models.Project | None = db.execute(
        select(models.Project).where(models.Project.id == project_id)
    ).scalar_one_or_none()

    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    if current_user.role != "admin" and project.owner_id != current_user.id:
        raise HTTPException(
            status_code=403, detail="Not authorized to access this project's files"
        )

    file: models.File | None = db.execute(
        select(models.File).where(
            models.File.project_id == project_id, models.File.id == file_id
        )
    ).scalar_one_or_none()

    if not file:
        raise HTTPException(status_code=404, detail="File not found")

    return schemas.File.model_validate(file)


@router.get("/{project_id}/file/{file_id}/content", response_class=Response)
def get_project_file_content(
    *,
    db: Session = Depends(get_db),
    project_id: int,
    file_id: int,
    preview: bool = Query(False),
    current_user: models.User = Depends(get_current_user),
) -> Response:
    """Retrieve the content of a file associated with a project."""
    project: models.Project | None = db.execute(
        select(models.Project).where(models.Project.id == project_id)
    ).scalar_one_or_none()

    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    if current_user.role != "admin" and project.owner_id != current_user.id:
        raise HTTPException(
            status_code=403, detail="Not authorized to access this project's files"
        )

    file: models.File | None = db.execute(
        select(models.File).where(
            models.File.project_id == project_id, models.File.id == file_id
        )
    ).scalar_one_or_none()

    if not file:
        raise HTTPException(status_code=404, detail="File not found")

    # Retrieve the file content from storage
    file_content = get_file(file.file_uuid)

    if preview:
        headers = {"Content-Disposition": f"inline; filename={file.file_name}"}
    else:
        headers = {"Content-Disposition": f"attachment; filename={file.file_name}"}
    return Response(content=file_content, media_type=file.file_type, headers=headers)


@router.post("/{project_id}/file", response_model=schemas.File)
def upload_file(
    *,
    db: Session = Depends(get_db),
    project_id: int,
    file: UploadFile = File(...),
    file_info: str = Form(...),
    current_user: models.User = Depends(get_current_user),
) -> schemas.File:
    """Upload a file with duplicate detection"""
    try:
        file_info = schemas.FileCreate.model_validate_json(file_info)
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Invalid JSON in file_info")
    except ValidationError as e:
        raise HTTPException(status_code=400, detail=str(e))

    check_project_access(project_id, current_user, db, permission="write")

    # Read file content
    file_content = file.file.read()
    file_size = len(file_content)
    file_hash = calculate_file_hash(file_content)

    # Check for duplicates
    existing_file = db.execute(
        select(models.File).where(
            and_(
                models.File.project_id == project_id, models.File.file_hash == file_hash
            )
        )
    ).scalar_one_or_none()

    if existing_file:
        raise HTTPException(
            status_code=409,
            detail={
                "message": "File already exists",
                "existing_file": {
                    "id": existing_file.id,
                    "file_name": existing_file.file_name,
                    "created_at": existing_file.created_at.isoformat(),
                },
            },
        )

    # Read MIME type from the uploaded file
    file_info.file_type = (
        file.content_type or file_info.file_type or "application/octet-stream"
    )

    # Save the file
    file_uuid = save_file(file_content)

    new_file = models.File(
        **file_info.model_dump(
            exclude={"file_uuid", "file_size", "file_hash", "file_metadata"}
        ),
        project_id=project_id,
        file_uuid=file_uuid,
        file_size=file_size,
        file_hash=file_hash,
        file_metadata=file_info.file_metadata
        if hasattr(file_info, "file_metadata")
        else None,
    )

    db.add(new_file)
    db.commit()
    db.refresh(new_file)
    return schemas.File.model_validate(new_file)


@router.post("/{project_id}/file/{file_id}/configure", response_model=schemas.File)
def configure_file_import(
    *,
    db: Session = Depends(get_db),
    project_id: int,
    file_id: int,
    config: dict = Body(...),
    current_user: models.User = Depends(get_current_user),
) -> schemas.File:
    """
    Set/update import config and preprocessing strategy for a file.
    """
    check_project_access(project_id, current_user, db, permission="write")

    file = db.execute(
        select(models.File).where(
            models.File.project_id == project_id, models.File.id == file_id
        )
    ).scalar_one_or_none()

    if not file:
        raise HTTPException(status_code=404, detail="File not found")

    # Accept "preprocessing_strategy" and/or "file_metadata"
    if "preprocessing_strategy" in config:
        file.preprocessing_strategy = config["preprocessing_strategy"]
    if "file_metadata" in config:
        file.file_metadata = config["file_metadata"]
    elif any(
        k in config
        for k in (
            "delimiter",
            "has_header",
            "row_split",
            "case_id_column",
            "text_columns",
        )
    ):
        file.file_metadata = {**(file.file_metadata or {}), **config}

    db.add(file)
    db.commit()
    db.refresh(file)

    return schemas.File.model_validate(file)


@router.get("/{project_id}/file/{file_id}/preview-rows", response_model=dict)
def preview_structured_file(
    *,
    db: Session = Depends(get_db),
    project_id: int,
    file_id: int,
    delimiter: str = Query(None, description="Delimiter for CSV"),
    encoding: str = Query("utf-8"),
    has_header: bool = Query(True),
    sheet: str = Query(None),
    max_rows: int = Query(10, ge=1, le=50),
    current_user: models.User = Depends(get_current_user),
) -> dict:
    """
    Return first N rows from CSV/XLSX for configuration/preview.
    """
    check_project_access(project_id, current_user, db)
    file = db.execute(
        select(models.File).where(
            models.File.project_id == project_id, models.File.id == file_id
        )
    ).scalar_one_or_none()
    if not file:
        raise HTTPException(status_code=404, detail="File not found")

    file_content = get_file(file.file_uuid)
    headers, rows, sheets = [], [], []
    if file.file_type == FileType.TEXT_CSV:
        # Try auto-detect if not specified
        sample_bytes = file_content[:16384]
        try:
            sample = sample_bytes.decode(encoding, errors="replace")
        except Exception:
            sample = sample_bytes.decode("utf-8", errors="replace")
        sniffer = csv.Sniffer()
        if not delimiter:
            try:
                dialect = sniffer.sniff(sample)
                delimiter = dialect.delimiter
            except Exception:
                delimiter = ","
        reader = csv.reader(io.StringIO(sample), delimiter=delimiter)
        all_rows = list(reader)
        if not all_rows:
            return {"headers": [], "rows": []}
        if has_header:
            headers = all_rows[0]
            rows = all_rows[1 : max_rows + 1]
        else:
            headers = [f"Column {i + 1}" for i in range(len(all_rows[0]))]
            rows = all_rows[:max_rows]
        return {
            "headers": headers,
            "rows": rows,
            "detected_delimiter": delimiter,
            "detected_encoding": encoding,
        }
    elif file.file_type in [
        FileType.APPLICATION_VND_MS_EXCEL,
        FileType.APPLICATION_VND_OPENXMLFORMATS_OFFICEDOCUMENT_SPREADSHEETML_SHEET,
    ]:
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True)
        if not sheet:
            sheet = wb.sheetnames[0]
        ws = wb[sheet]
        sheets = wb.sheetnames
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            rows.append(list(row))
            if i + 1 >= max_rows + (1 if has_header else 0):
                break
        if has_header and rows:
            headers = list(rows[0])
            rows = rows[1:]
        else:
            headers = [f"Column {i + 1}" for i in range(len(rows[0]))]
        return {"headers": headers, "rows": rows, "sheets": sheets}
    else:
        raise HTTPException(
            status_code=400, detail="Preview not supported for this file type"
        )


@router.post("/{project_id}/file/check-duplicates", response_model=list[dict])
def check_duplicates(
    *,
    db: Session = Depends(get_db),
    project_id: int,
    files: list[dict] = Body(..., description="List of {filename, hash} objects"),
    current_user: models.User = Depends(get_current_user),
) -> list[dict]:
    """Check for duplicate files before uploading"""
    check_project_access(project_id, current_user, db)

    results = []
    for file_info in files:
        existing = db.execute(
            select(models.File).where(
                and_(
                    models.File.project_id == project_id,
                    models.File.file_hash == file_info["hash"],
                )
            )
        ).scalar_one_or_none()

        results.append(
            {
                "filename": file_info["filename"],
                "hash": file_info["hash"],
                "exists": existing is not None,
                "existing_file": {
                    "id": existing.id,
                    "file_name": existing.file_name,
                    "created_at": existing.created_at.isoformat(),
                }
                if existing
                else None,
            }
        )

    return results


@router.delete("/{project_id}/file/{file_id}", response_model=schemas.File)
def delete_file(
    *,
    db: Session = Depends(get_db),
    project_id: int,
    file_id: int,
    force: bool = Query(False, description="Force delete even if linked"),
    current_user: models.User = Depends(get_current_user),
) -> schemas.File:
    """Delete a file with safety checks"""
    check_project_access(project_id, current_user, db, permission="write")

    file = db.execute(
        select(models.File)
        .options(
            selectinload(models.File.documents_as_original),
            selectinload(models.File.documents_as_preprocessed),
            selectinload(models.File.preprocessing_tasks),
            selectinload(models.File.file_preprocessing_tasks),
        )
        .where(models.File.project_id == project_id, models.File.id == file_id)
    ).scalar_one_or_none()

    if not file:
        raise HTTPException(status_code=404, detail="File not found")

    # Check if file is linked to other resources
    is_linked = (
        len(file.documents_as_original) > 0
        or len(file.documents_as_preprocessed) > 0
        or len(file.preprocessing_tasks) > 0
        or len(file.file_preprocessing_tasks) > 0
    )

    if is_linked and not force:
        raise HTTPException(
            status_code=409,
            detail={
                "message": "File is linked to other resources",
                "links": {
                    "documents": len(file.documents_as_original)
                    + len(file.documents_as_preprocessed),
                    "preprocessing_tasks": len(file.preprocessing_tasks)
                    + len(file.file_preprocessing_tasks),
                },
            },
        )

    # Delete the file content from storage
    try:
        remove_file(file.file_uuid)
    except FileNotFoundError:
        # Log the error but continue with database deletion
        pass

    db.delete(file)
    db.commit()

    return schemas.File.model_validate(file)


@router.post("/{project_id}/file/batch-delete", response_model=dict)
def batch_delete_files(
    *,
    db: Session = Depends(get_db),
    project_id: int,
    file_ids: list[int] = Body(...),
    force: bool = Body(False),
    current_user: models.User = Depends(get_current_user),
) -> dict:
    """Delete multiple files at once"""
    check_project_access(project_id, current_user, db, permission="write")

    deleted = []
    errors = []

    for file_id in file_ids:
        try:
            _ = delete_file(
                db=db,
                project_id=project_id,
                file_id=file_id,
                force=force,
                current_user=current_user,
            )
            deleted.append(file_id)
        except HTTPException as e:
            errors.append({"file_id": file_id, "error": e.detail})
        except Exception as e:
            errors.append({"file_id": file_id, "error": str(e)})

    return {
        "deleted": deleted,
        "errors": errors,
        "total_deleted": len(deleted),
        "total_errors": len(errors),
    }


# Continue the endpoint in project.py


@router.post("/{project_id}/file/download-zip")
async def download_files_as_zip(
    *,
    db: Session = Depends(get_db),
    project_id: int,
    file_ids: list[int] = Body(...),
    include_metadata: bool = Body(True),
    current_user: models.User = Depends(get_current_user),
) -> Response:
    """Download multiple files as a ZIP archive"""
    check_project_access(project_id, current_user, db)

    import io
    import json
    import zipfile

    # Create ZIP file in memory
    zip_buffer = io.BytesIO()

    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
        files_metadata = []

        for file_id in file_ids:
            file = db.execute(
                select(models.File).where(
                    models.File.id == file_id, models.File.project_id == project_id
                )
            ).scalar_one_or_none()

            if not file:
                continue

            try:
                # Get file content
                file_content = get_file(file.file_uuid)

                # Add file to ZIP
                zip_file.writestr(file.file_name, file_content)

                # Collect metadata
                if include_metadata:
                    files_metadata.append(
                        {
                            "id": file.id,
                            "file_name": file.file_name,
                            "file_type": file.file_type,
                            "file_size": file.file_size,
                            "file_hash": file.file_hash,
                            "description": file.description,
                            "created_at": file.created_at.isoformat()
                            if file.created_at
                            else None,
                            "updated_at": file.updated_at.isoformat()
                            if file.updated_at
                            else None,
                        }
                    )

            except Exception as e:
                print(f"Error adding file {file.file_name} to ZIP: {e}")
                continue

        # Add metadata file if requested
        if include_metadata and files_metadata:
            metadata_json = json.dumps(
                {
                    "project_id": project_id,
                    "export_date": datetime.datetime.now(datetime.UTC).isoformat(),
                    "total_files": len(files_metadata),
                    "files": files_metadata,
                },
                indent=2,
            )
            zip_file.writestr("metadata.json", metadata_json)

    # Prepare response
    zip_buffer.seek(0)
    filename = f"project_{project_id}_files_{datetime.datetime.now(datetime.UTC).strftime('%Y%m%d_%H%M%S')}.zip"

    return Response(
        content=zip_buffer.getvalue(),
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/{project_id}/file/move")
async def move_files(
    *,
    db: Session = Depends(get_db),
    project_id: int,
    file_ids: list[int] = Body(...),
    target_project_id: int = Body(...),
    current_user: models.User = Depends(get_current_user),
) -> dict:
    """Move files to another project"""
    # Check access to both source and target projects
    check_project_access(project_id, current_user, db, permission="write")
    check_project_access(target_project_id, current_user, db, permission="write")

    if project_id == target_project_id:
        raise HTTPException(
            status_code=400, detail="Source and target projects cannot be the same"
        )

    moved_count = 0
    errors = []

    for file_id in file_ids:
        try:
            file = db.execute(
                select(models.File).where(
                    models.File.id == file_id, models.File.project_id == project_id
                )
            ).scalar_one_or_none()

            if not file:
                errors.append({"file_id": file_id, "error": "File not found"})
                continue

            # Update the file's project
            file.project_id = target_project_id

            # Update any related documents
            documents = (
                db.execute(
                    select(models.Document).where(
                        or_(
                            models.Document.original_file_id == file_id,
                            models.Document.preprocessed_file_id == file_id,
                        )
                    )
                )
                .scalars()
                .all()
            )

            for doc in documents:
                doc.project_id = target_project_id

            moved_count += 1

        except Exception as e:
            errors.append({"file_id": file_id, "error": str(e)})

    db.commit()

    return {"moved": moved_count, "errors": errors, "total_requested": len(file_ids)}


@router.get("/{project_id}/file/check-links/{file_id}")
def check_file_links(
    *,
    db: Session = Depends(get_db),
    project_id: int,
    file_id: int,
    current_user: models.User = Depends(get_current_user),
) -> dict:
    """Check if a file has any linked resources"""
    check_project_access(project_id, current_user, db)

    file = db.execute(
        select(models.File)
        .options(
            selectinload(models.File.documents_as_original),
            selectinload(models.File.documents_as_preprocessed),
            selectinload(models.File.preprocessing_tasks),
            selectinload(models.File.file_preprocessing_tasks),
        )
        .where(models.File.project_id == project_id, models.File.id == file_id)
    ).scalar_one_or_none()

    if not file:
        raise HTTPException(status_code=404, detail="File not found")

    return {
        "file_id": file_id,
        "is_linked": bool(
            file.documents_as_original
            or file.documents_as_preprocessed
            or file.preprocessing_tasks
            or file.file_preprocessing_tasks
        ),
        "links": {
            "documents_as_original": len(file.documents_as_original),
            "documents_as_preprocessed": len(file.documents_as_preprocessed),
            "preprocessing_tasks": len(file.preprocessing_tasks),
            "file_preprocessing_tasks": len(file.file_preprocessing_tasks),
        },
    }


@router.get(
    "/{project_id}/preprocessing-config",
    response_model=List[schemas.PreprocessingConfiguration],
)
def get_preprocessing_configurations(
    *,
    db: Session = Depends(get_db),
    project_id: int,
    file_type: str | None = None,
    current_user: models.User = Depends(get_current_user),
) -> List[schemas.PreprocessingConfiguration]:
    """Get all preprocessing configurations for a project."""
    check_project_access(project_id, current_user, db, "read")

    query = select(models.PreprocessingConfiguration).where(
        models.PreprocessingConfiguration.project_id == project_id
    )

    if file_type:
        file_type_enum = models.FileType(file_type)  # now accepts "mixed"
        query = query.where(
            models.PreprocessingConfiguration.file_type == file_type_enum
        )

    configs = db.execute(query).scalars().all()
    return [schemas.PreprocessingConfiguration.model_validate(c) for c in configs]


@router.get(
    "/{project_id}/preprocessing-config/{config_id}",
    response_model=schemas.PreprocessingConfiguration,
)
def get_preprocessing_configuration(
    *,
    db: Session = Depends(get_db),
    project_id: int,
    config_id: int,
    current_user: models.User = Depends(get_current_user),
) -> schemas.PreprocessingConfiguration:
    """Get a specific preprocessing configuration."""
    check_project_access(project_id, current_user, db, "read")

    config = db.execute(
        select(models.PreprocessingConfiguration).where(
            models.PreprocessingConfiguration.id == config_id,
            models.PreprocessingConfiguration.project_id == project_id,
        )
    ).scalar_one_or_none()

    if not config:
        raise HTTPException(status_code=404, detail="Configuration not found")

    return schemas.PreprocessingConfiguration.model_validate(config)


@router.post(
    "/{project_id}/preprocessing-config",
    response_model=schemas.PreprocessingConfiguration,
)
def create_preprocessing_configuration(
    *,
    db: Session = Depends(get_db),
    project_id: int,
    config: schemas.PreprocessingConfigurationCreate,
    current_user: models.User = Depends(get_current_user),
) -> schemas.PreprocessingConfiguration:
    """Create a reusable preprocessing configuration."""
    check_project_access(project_id, current_user, db, "write")

    # Check for duplicate names
    existing = db.execute(
        select(models.PreprocessingConfiguration).where(
            models.PreprocessingConfiguration.project_id == project_id,
            models.PreprocessingConfiguration.name == config.name,
        )
    ).scalar_one_or_none()

    if existing:
        raise HTTPException(
            status_code=400, detail="Configuration with this name already exists"
        )

    # Create new configuration
    db_config = models.PreprocessingConfiguration(
        **config.model_dump(), project_id=project_id
    )
    db.add(db_config)
    db.commit()
    db.refresh(db_config)

    return schemas.PreprocessingConfiguration.model_validate(db_config)


@router.put(
    "/{project_id}/preprocessing-config/{config_id}",
    response_model=schemas.PreprocessingConfiguration,
)
def update_preprocessing_configuration(
    *,
    db: Session = Depends(get_db),
    project_id: int,
    config_id: int,
    config_update: schemas.PreprocessingConfigurationUpdate,
    current_user: models.User = Depends(get_current_user),
) -> schemas.PreprocessingConfiguration:
    """Update a preprocessing configuration."""
    check_project_access(project_id, current_user, db, "write")

    config = db.execute(
        select(models.PreprocessingConfiguration).where(
            models.PreprocessingConfiguration.id == config_id,
            models.PreprocessingConfiguration.project_id == project_id,
        )
    ).scalar_one_or_none()

    if not config:
        raise HTTPException(status_code=404, detail="Configuration not found")

    # Check if configuration is in use by active task
    active_tasks = (
        db.execute(
            select(models.PreprocessingTask).where(
                models.PreprocessingTask.configuration_id == config_id,
                models.PreprocessingTask.status.in_(
                    [
                        models.PreprocessingStatus.PENDING,
                        models.PreprocessingStatus.IN_PROGRESS,
                    ]
                ),
            )
        )
        .scalars()
        .first()
    )

    if active_tasks:
        raise HTTPException(
            status_code=400,
            detail="Cannot update configuration while it's being used in active tasks",
        )

    # Is config referenced by any document?
    in_use = (
        db.query(models.Document)
        .filter(models.Document.preprocessing_config_id == config.id)
        .count()
        > 0
    )

    allowed_fields = {"name", "description"}
    changes = config_update.model_dump(exclude_unset=True)

    if in_use:
        # Find which *disallowed* fields are *actually being changed*
        disallowed_changed = [
            field
            for field, value in changes.items()
            if field not in allowed_fields and getattr(config, field) != value
        ]
        if disallowed_changed:
            raise HTTPException(
                status_code=400,
                detail=(
                    "Cannot edit a configuration in use by existing documents "
                    "(except name/description). The following fields would change: "
                    f"{', '.join(disallowed_changed)}"
                ),
            )

    # Apply the changes
    for field, value in changes.items():
        setattr(config, field, value)

    db.commit()
    db.refresh(config)
    return schemas.PreprocessingConfiguration.model_validate(config)


@router.delete("/{project_id}/preprocessing-config/{config_id}")
def delete_preprocessing_configuration(
    *,
    db: Session = Depends(get_db),
    project_id: int,
    config_id: int,
    current_user: models.User = Depends(get_current_user),
) -> dict:
    """Delete a preprocessing configuration."""
    check_project_access(project_id, current_user, db, "write")

    config = db.execute(
        select(models.PreprocessingConfiguration).where(
            models.PreprocessingConfiguration.id == config_id,
            models.PreprocessingConfiguration.project_id == project_id,
        )
    ).scalar_one_or_none()

    if not config:
        raise HTTPException(status_code=404, detail="Configuration not found")

    # Check if configuration has been used
    tasks_using_config = (
        db.execute(
            select(models.PreprocessingTask).where(
                models.PreprocessingTask.configuration_id == config_id
            )
        )
        .scalars()
        .first()
    )

    if tasks_using_config:
        raise HTTPException(
            status_code=400,
            detail="Cannot delete configuration that has been used in preprocessing tasks",
        )

    db.delete(config)
    db.commit()

    return {"detail": "Configuration deleted successfully"}


@router.post("/{project_id}/preprocess", response_model=schemas.PreprocessingTask)
async def preprocess_project_data(
    *,
    project_id: int,
    preprocessing_task: schemas.PreprocessingTaskCreate,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> schemas.PreprocessingTask:
    """Start preprocessing with advanced duplicate detection and progress tracking."""
    check_project_access(project_id, current_user, db, "write")

    # Determine configuration to use
    if preprocessing_task.configuration_id:
        # Using existing configuration
        config = db.get(
            models.PreprocessingConfiguration, preprocessing_task.configuration_id
        )
        if not config or config.project_id != project_id:
            raise HTTPException(status_code=404, detail="Configuration not found")
    elif preprocessing_task.inline_config:
        # Check if a matching configuration already exists
        config_dict = preprocessing_task.inline_config.model_dump(
            exclude={"bypass_celery"}
        )

        # Build query to find matching configuration
        query = db.query(models.PreprocessingConfiguration).filter(
            models.PreprocessingConfiguration.project_id == project_id,
            models.PreprocessingConfiguration.pdf_backend
            == config_dict.get("pdf_backend"),
            models.PreprocessingConfiguration.ocr_backend
            == config_dict.get("ocr_backend"),
            models.PreprocessingConfiguration.use_ocr
            == config_dict.get("use_ocr", True),
            models.PreprocessingConfiguration.force_ocr
            == config_dict.get("force_ocr", False),
        )

        # Handle None values for ocr_model
        if config_dict.get("ocr_model") is None:
            query = query.filter(models.PreprocessingConfiguration.ocr_model.is_(None))
        else:
            query = query.filter(
                models.PreprocessingConfiguration.ocr_model
                == config_dict.get("ocr_model")
            )

        # Get all potential matches to check complex fields
        potential_matches = query.all()

        # Check OCR languages and other complex fields
        existing_config = None
        for potential_config in potential_matches:
            # Compare OCR languages (handle None and empty lists)
            config_langs = sorted(potential_config.ocr_languages or [])
            new_langs = sorted(config_dict.get("ocr_languages") or [])

            if config_langs != new_langs:
                continue

            # Compare additional_settings (handle None)
            if (potential_config.additional_settings or {}) != (
                config_dict.get("additional_settings") or {}
            ):
                continue

            # For "Quick Process" or similar standard configs, also check the name
            if config_dict.get("name") in [
                "Quick Process",
                "Custom Process",
                "Standard Process",
            ]:
                if potential_config.name == config_dict.get("name"):
                    existing_config = potential_config
                    break
            else:
                # For other configs, any match is good
                existing_config = potential_config
                break

        if existing_config:
            # Use existing configuration
            config = existing_config
        else:
            # Create new configuration only if none exists
            config = models.PreprocessingConfiguration(
                project_id=project_id,
                name=config_dict.get(
                    "name", f"Auto-created config {datetime.datetime.now(datetime.UTC)}"
                ),
                description=config_dict.get(
                    "description", "Automatically created configuration"
                ),
                pdf_backend=config_dict.get("pdf_backend"),
                ocr_backend=config_dict.get("ocr_backend"),
                use_ocr=config_dict.get("use_ocr", True),
                force_ocr=config_dict.get("force_ocr", False),
                ocr_languages=config_dict.get("ocr_languages"),
                ocr_model=config_dict.get("ocr_model"),
                llm_model=config_dict.get("llm_model"),
                additional_settings=config_dict.get("additional_settings"),
            )
            db.add(config)
            db.commit()
            db.refresh(config)
    else:
        raise HTTPException(
            status_code=400,
            detail="Either configuration_id or inline_config must be provided",
        )

    # Validate files exist and belong to project
    files = (
        db.execute(
            select(models.File).where(
                models.File.id.in_(preprocessing_task.file_ids),
                models.File.project_id == project_id,
            )
        )
        .scalars()
        .all()
    )
    if len(files) != len(preprocessing_task.file_ids):
        missing_ids = set(preprocessing_task.file_ids) - {f.id for f in files}
        raise HTTPException(
            status_code=404,
            detail=f"Files not found or don't belong to project: {missing_ids}",
        )

    # Create preprocessing task
    task = models.PreprocessingTask(
        project_id=project_id,
        configuration_id=config.id,
        total_files=len(files),
        rollback_on_cancel=preprocessing_task.rollback_on_cancel,
    )

    # Store API credentials in task metadata if provided
    if preprocessing_task.api_key and preprocessing_task.base_url:
        task.task_metadata = {
            "custom_api_used": True,
            "api_base_url": preprocessing_task.base_url,
            # Don't store the actual API key for security
        }

    db.add(task)
    db.commit()

    # Check for duplicates and create file tasks
    file_tasks_to_process = []
    skipped_files = 0
    skipped_file_names = []

    for file in files:
        # Check for existing documents with same configuration
        existing_docs = (
            db.execute(
                select(models.Document).where(
                    models.Document.original_file_id == file.id,
                    models.Document.preprocessing_config_id == config.id,
                )
            )
            .scalars()
            .all()
        )

        if existing_docs and not preprocessing_task.force_reprocess:
            # Skip this file
            task.processed_files += 1
            skipped_files += 1
            skipped_file_names.append(file.file_name)
            continue

        # Delete existing documents if force reprocess
        if existing_docs and preprocessing_task.force_reprocess:
            for doc in existing_docs:
                doc.document_sets.clear()
                db.delete(doc)

        file_task = models.FilePreprocessingTask(
            preprocessing_task_id=task.id,
            file_id=file.id,
            file_name=file.file_name,  # Set file name immediately
        )
        db.add(file_task)
        file_tasks_to_process.append(file_task)

    db.commit()

    # Update task with skipped files information
    if skipped_files > 0:
        if not task.task_metadata:
            task.task_metadata = {}
        task.task_metadata["skipped_files"] = skipped_files
        task.task_metadata["skipped_file_names"] = skipped_file_names
        task.skipped_files = skipped_files

    if not file_tasks_to_process:
        task.status = models.PreprocessingStatus.COMPLETED
        task.message = f"All files already processed with these settings. {skipped_files} files skipped."
        task.completed_at = datetime.datetime.now(datetime.UTC)
        db.commit()
        db.refresh(task)
        return schemas.PreprocessingTask.model_validate(task)

    # Set initial message
    if skipped_files > 0:
        task.message = f"Processing {len(file_tasks_to_process)} files. {skipped_files} files already processed and skipped."

    bypass_celery = getattr(preprocessing_task, "bypass_celery", False)
    inline_cfg = getattr(preprocessing_task, "inline_config", None)
    if inline_cfg:
        if hasattr(inline_cfg, "bypass_celery"):
            bypass_celery = getattr(inline_cfg, "bypass_celery", False)
        elif isinstance(inline_cfg, dict):
            bypass_celery = inline_cfg.get("bypass_celery", False)

    # 2. Only check admin if someone tried to set bypass_celery
    if bypass_celery:
        if not await get_admin_user(current_user):
            # Not an admin, so do not allow bypass
            bypass_celery = False
            # Optionally, raise error instead:
            raise HTTPException(403, "Only admins may set bypass_celery")

    # Start processing
    if bypass_celery:
        print("Bypassing Celery for preprocessing task")
        from ....utils.preprocessing import PreprocessingPipeline

        try:
            # Pass API credentials to pipeline
            pipeline = PreprocessingPipeline(
                db,
                task.id,
                api_key=preprocessing_task.api_key,
                base_url=preprocessing_task.base_url,
            )
            pipeline.process()
        except Exception as e:
            task.status = models.PreprocessingStatus.FAILED
            task.message = f"Processing failed: {str(e)}"
            db.commit()
            raise HTTPException(status_code=500, detail=str(e))
    else:
        from ....celery.preprocessing import process_files_async

        # Pass credentials through celery
        result = process_files_async.delay(
            task.id,
            api_key=preprocessing_task.api_key,
            base_url=preprocessing_task.base_url,
        )
        task.celery_task_id = result.id
        db.commit()

    db.commit()
    db.refresh(task)
    return schemas.PreprocessingTask.model_validate(task)


@router.get("/{project_id}/preprocess", response_model=List[schemas.PreprocessingTask])
def get_preprocessing_tasks(
    *,
    project_id: int,
    status: str | None = Query(None, description="Filter by status"),
    limit: int = Query(100, ge=1, le=1000),
    offset: int = Query(0, ge=0),
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> List[schemas.PreprocessingTask]:
    """Get preprocessing tasks for a project."""
    check_project_access(project_id, current_user, db, "read")

    query = (
        select(models.PreprocessingTask)
        .where(models.PreprocessingTask.project_id == project_id)
        .options(
            selectinload(models.PreprocessingTask.file_tasks).selectinload(
                models.FilePreprocessingTask.file
            ),
            selectinload(models.PreprocessingTask.configuration),
        )
    )

    if status:
        try:
            status_enum = models.PreprocessingStatus(status)
            query = query.where(models.PreprocessingTask.status == status_enum)
        except ValueError:
            raise HTTPException(status_code=400, detail=f"Invalid status: {status}")

    query = query.order_by(models.PreprocessingTask.created_at.desc())
    query = query.limit(limit).offset(offset)

    tasks = db.execute(query).scalars().all()
    return [schemas.PreprocessingTask.model_validate(task) for task in tasks]


@router.get(
    "/{project_id}/preprocess/{task_id}", response_model=schemas.PreprocessingTask
)
def get_preprocessing_task(
    *,
    project_id: int,
    task_id: int,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> schemas.PreprocessingTask:
    """Get detailed information about a preprocessing task."""
    check_project_access(project_id, current_user, db, "read")

    task = db.execute(
        select(models.PreprocessingTask)
        .where(
            models.PreprocessingTask.id == task_id,
            models.PreprocessingTask.project_id == project_id,
        )
        .options(
            selectinload(models.PreprocessingTask.file_tasks).selectinload(
                models.FilePreprocessingTask.file
            ),
            selectinload(models.PreprocessingTask.configuration),
        )
    ).scalar_one_or_none()

    if not task:
        raise HTTPException(status_code=404, detail="Task not found")

    return schemas.PreprocessingTask.model_validate(task)


@router.post(
    "/{project_id}/preprocess/{task_id}/cancel",
    response_model=schemas.PreprocessingTask,
)
def cancel_preprocessing_task(
    *,
    project_id: int,
    task_id: int,
    keep_processed: bool = Query(False, description="Keep already processed files"),
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> schemas.PreprocessingTask:
    """Cancel a preprocessing task with option to keep or rollback processed files."""

    check_project_access(project_id, current_user, db, "write")

    task = (
        db.query(models.PreprocessingTask)
        .filter(
            models.PreprocessingTask.id == task_id,
            models.PreprocessingTask.project_id == project_id,
        )
        .first()
    )

    if not task:
        raise HTTPException(status_code=404, detail="Task not found")

    if task.status in [
        models.PreprocessingStatus.COMPLETED,
        models.PreprocessingStatus.CANCELLED,
    ]:
        raise HTTPException(
            status_code=400, detail=f"Cannot cancel task in {task.status} status"
        )

    # Set cancellation flag
    task.is_cancelled = True
    task.status = models.PreprocessingStatus.CANCELLED
    task.completed_at = datetime.datetime.now(datetime.UTC)

    # Rollback logic: remove processed docs if requested
    if not keep_processed and task.rollback_on_cancel:
        deleted_count = 0
        for file_task in task.file_tasks:
            if file_task.status == models.PreprocessingStatus.COMPLETED:
                for doc in file_task.documents:
                    doc.document_sets.clear()
                    db.delete(doc)
                    deleted_count += 1
        task.message = (
            f"Task cancelled and {deleted_count} processed documents rolled back"
        )
    else:
        task.message = "Task cancelled, keeping processed documents"

    # Mark all still-pending/in-progress file tasks as cancelled
    for file_task in task.file_tasks:
        if file_task.status in [
            models.PreprocessingStatus.PENDING,
            models.PreprocessingStatus.IN_PROGRESS,
        ]:
            file_task.status = models.PreprocessingStatus.CANCELLED
            file_task.completed_at = datetime.datetime.now(datetime.UTC)

    db.commit()
    db.refresh(task)

    return schemas.PreprocessingTask.model_validate(task)


@router.get("/{project_id}/preprocess/{task_id}/progress")
def get_preprocessing_progress(
    *,
    project_id: int,
    task_id: int,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> schemas.PreprocessingTask:
    """Get detailed progress of a preprocessing task."""
    check_project_access(project_id, current_user, db, "read")

    task = db.execute(
        select(models.PreprocessingTask).where(
            models.PreprocessingTask.id == task_id,
            models.PreprocessingTask.project_id == project_id,
        )
    ).scalar_one_or_none()

    if not task:
        raise HTTPException(status_code=404, detail="Task not found")

    # Calculate estimated completion time if in progress
    if (
        task.status == models.PreprocessingStatus.IN_PROGRESS
        and task.processed_files > 0
    ):
        elapsed = datetime.datetime.now(datetime.UTC) - task.started_at
        avg_time_per_file = elapsed.total_seconds() / task.processed_files
        remaining_files = task.total_files - task.processed_files - task.failed_files

        if remaining_files > 0:
            estimated_remaining = datetime.timedelta(
                seconds=avg_time_per_file * remaining_files
            )
            task.estimated_completion = (
                datetime.datetime.now(datetime.UTC) + estimated_remaining
            )

    return schemas.PreprocessingTask.model_validate(task)


@router.get("/{project_id}/preprocess/{task_id}/retry-failed")
def retry_failed_files(
    *,
    project_id: int,
    task_id: int,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> schemas.PreprocessingTask:
    """Create a new task to retry failed files from a previous task."""
    check_project_access(project_id, current_user, db, "write")

    original_task = db.execute(
        select(models.PreprocessingTask).where(
            models.PreprocessingTask.id == task_id,
            models.PreprocessingTask.project_id == project_id,
        )
    ).scalar_one_or_none()

    if not original_task:
        raise HTTPException(status_code=404, detail="Task not found")

    # Get failed file IDs
    failed_file_ids = [
        ft.file_id
        for ft in original_task.file_tasks
        if ft.status == models.PreprocessingStatus.FAILED
    ]

    if not failed_file_ids:
        raise HTTPException(status_code=400, detail="No failed files to retry")

    # Create new task with same configuration
    new_task = models.PreprocessingTask(
        project_id=project_id,
        configuration_id=original_task.configuration_id,
        total_files=len(failed_file_ids),
        rollback_on_cancel=original_task.rollback_on_cancel,
    )
    db.add(new_task)
    db.commit()

    # Create file tasks for failed files
    for file_id in failed_file_ids:
        file_task = models.FilePreprocessingTask(
            preprocessing_task_id=new_task.id, file_id=file_id
        )
        db.add(file_task)

    db.commit()

    # Start processing
    from ....celery.preprocessing import process_files_async

    result = process_files_async.delay(new_task.id)
    new_task.celery_task_id = result.id
    db.commit()

    db.refresh(new_task)
    return schemas.PreprocessingTask.model_validate(new_task)


@router.get("/{project_id}/document", response_model=List[schemas.Document])
def get_documents(
    *,
    project_id: int,
    file_id: str | None = Query(None, description="Filter by original file"),
    preprocessing_task_id: int | None = Query(
        None, description="Filter by preprocessing task"
    ),
    limit: int = Query(100, ge=1, le=1000),
    offset: int = Query(0, ge=0),
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> List[schemas.Document]:
    """Get documents for a project with optional filtering."""
    check_project_access(project_id, current_user, db, "read")

    query = select(models.Document).where(models.Document.project_id == project_id)

    if file_id:
        query = query.where(models.Document.original_file_id == file_id)

    if preprocessing_task_id:
        query = query.join(models.FilePreprocessingTask).where(
            models.FilePreprocessingTask.preprocessing_task_id == preprocessing_task_id
        )

    query = query.order_by(models.Document.created_at.desc())
    query = query.limit(limit).offset(offset)

    documents = db.execute(query).scalars().all()
    return [schemas.Document.model_validate(doc) for doc in documents]


@router.get("/{project_id}/document/{document_id}", response_model=schemas.Document)
def get_document(
    *,
    project_id: int,
    document_id: int,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> schemas.Document:
    project: models.Project | None = db.execute(
        select(models.Project).where(models.Project.id == project_id)
    ).scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    if current_user.role != "admin" and project.owner_id != current_user.id:
        raise HTTPException(
            status_code=403, detail="Not authorized to access this project's documents"
        )

    document: models.Document | None = db.execute(
        select(models.Document).where(
            models.Document.project_id == project_id,
            models.Document.id == document_id,
        )
    ).scalar_one_or_none()
    if not document:
        raise HTTPException(status_code=404, detail="Document not found")

    return schemas.Document.model_validate(document)


@router.delete("/{project_id}/document/{document_id}")
def delete_document(
    *,
    project_id: int,
    document_id: int,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    """Delete a specific document, only if not used in any trial, trial result, or evaluation metric."""
    check_project_access(project_id, current_user, db, "write")

    document = db.execute(
        select(models.Document).where(
            models.Document.id == document_id, models.Document.project_id == project_id
        )
    ).scalar_one_or_none()

    if not document:
        raise HTTPException(status_code=404, detail="Document not found")

    # --- NEW: Check for usage in any trial (document_ids is a JSON list) ---
    trials_with_doc = (
        db.execute(
            select(models.Trial).where(
                models.Trial.project_id == project_id,
                models.Trial.document_ids.contains([document_id]),
            )
        )
        .scalars()
        .first()
    )
    if trials_with_doc:
        raise HTTPException(
            status_code=400,
            detail=f"Document is referenced in trial '{trials_with_doc.name or trials_with_doc.id}'. Remove from trial(s) first.",
        )

    # --- Check if document is used in any trial results ---
    trial_result = db.execute(
        select(models.TrialResult).where(models.TrialResult.document_id == document_id)
    ).scalar_one_or_none()
    if trial_result:
        raise HTTPException(
            status_code=400,
            detail="Document is referenced in a trial result. Remove results/trials first.",
        )

    # --- (Optional) Check if document is used in any evaluation metric ---
    metric = db.execute(
        select(models.EvaluationMetric).where(
            models.EvaluationMetric.document_id == document_id
        )
    ).scalar_one_or_none()
    if metric:
        raise HTTPException(
            status_code=400,
            detail="Document is referenced in evaluation metrics. Remove related evaluation/trial first.",
        )

    # --- Existing check: Document sets ---
    if document.document_sets:
        raise HTTPException(
            status_code=400,
            detail=f"Document is part of {len(document.document_sets)} document sets. Remove from sets first.",
        )

    # --- Preprocessed file deletion logic as before ---
    if document.preprocessed_file_id:
        other_docs_using_file = db.execute(
            select(models.Document).where(
                models.Document.preprocessed_file_id == document.preprocessed_file_id,
                models.Document.id != document_id,
            )
        ).scalar_one_or_none()

        if not other_docs_using_file:
            preprocessed_file = db.get(models.File, document.preprocessed_file_id)
            if preprocessed_file:
                try:
                    from ....dependencies import remove_file

                    remove_file(preprocessed_file.file_uuid)
                    db.delete(preprocessed_file)
                except Exception as e:
                    print(f"Error deleting preprocessed file: {e}")

    db.delete(document)
    db.commit()

    return {"detail": "Document deleted successfully"}


@router.post("/{project_id}/document-set", response_model=schemas.DocumentSet)
def create_document_set(
    project_id: int,
    document_set: schemas.DocumentSetCreate,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> schemas.DocumentSet:
    """Create a document set from documents or a trial"""
    # Check project access
    project: models.Project | None = db.execute(
        select(models.Project).where(models.Project.id == project_id)
    ).scalar_one_or_none()

    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    if current_user.role != "admin" and project.owner_id != current_user.id:
        raise HTTPException(
            status_code=403,
            detail="Not authorized to create document sets for this project",
        )

    # Create the document set
    db_set = models.DocumentSet(
        project_id=project_id,
        name=document_set.name,
        description=document_set.description,
    )

    db.add(db_set)
    db.flush()

    # Add documents based on whether it's from trial or direct selection
    if document_set.trial_id:
        # Get documents from trial
        trial: models.Trial | None = db.execute(
            select(models.Trial).where(
                models.Trial.id == document_set.trial_id,
                models.Trial.project_id == project_id,
            )
        ).scalar_one_or_none()

        if not trial:
            raise HTTPException(
                status_code=404, detail="Trial not found in this project"
            )

        document_ids = trial.document_ids
    else:
        # Use provided document IDs
        document_ids = document_set.document_ids

    # Add documents to the set
    for doc_id in document_ids:
        # Verify document exists in project
        doc = db.execute(
            select(models.Document).where(
                models.Document.id == doc_id, models.Document.project_id == project_id
            )
        ).scalar_one_or_none()

        if doc:
            db.execute(
                document_set_association.insert().values(
                    document_id=doc_id, document_set_id=db_set.id
                )
            )

    db.commit()
    db.refresh(db_set)

    return schemas.DocumentSet.model_validate(db_set)


@router.get("/{project_id}/document-set", response_model=List[schemas.DocumentSet])
def get_document_sets(
    project_id: int,
    include_auto_generated: bool = Query(
        False, description="Include auto-generated sets from trials"
    ),
    preprocessing_config_id: int = Query(
        None, description="Filter by preprocessing configuration"
    ),
    tag: str = Query(None, description="Filter by tag"),
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> List[schemas.DocumentSet]:
    check_project_access(project_id, current_user, db, "read")

    query = select(models.DocumentSet).where(
        models.DocumentSet.project_id == project_id
    )

    if not include_auto_generated:
        query = query.where(~models.DocumentSet.is_auto_generated)

    if preprocessing_config_id:
        query = query.where(
            models.DocumentSet.preprocessing_config_id == preprocessing_config_id
        )

    if tag:
        query = query.where(models.DocumentSet.tags.contains([tag]))

    sets = (
        db.execute(query.order_by(models.DocumentSet.created_at.desc())).scalars().all()
    )

    return [schemas.DocumentSet.model_validate(s) for s in sets]


@router.patch("/{project_id}/document-set/{set_id}", response_model=schemas.DocumentSet)
def update_document_set(
    project_id: int,
    set_id: int,
    update_data: schemas.DocumentSetUpdate,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> schemas.DocumentSet:
    check_project_access(project_id, current_user, db, "write")

    doc_set = db.execute(
        select(models.DocumentSet).where(
            models.DocumentSet.id == set_id, models.DocumentSet.project_id == project_id
        )
    ).scalar_one_or_none()

    if not doc_set:
        raise HTTPException(status_code=404, detail="Document set not found")

    # Update fields
    for field, value in update_data.model_dump(exclude_unset=True).items():
        if field == "document_ids":
            # Handle document updates
            # Remove existing associations
            db.execute(
                document_set_association.delete().where(
                    document_set_association.c.document_set_id == set_id
                )
            )
            # Add new associations
            for doc_id in value:
                db.execute(
                    document_set_association.insert().values(
                        document_id=doc_id, document_set_id=set_id
                    )
                )
        else:
            setattr(doc_set, field, value)

    db.commit()
    db.refresh(doc_set)

    return schemas.DocumentSet.model_validate(doc_set)


@router.delete(
    "/{project_id}/document-set/{set_id}",
    status_code=status.HTTP_204_NO_CONTENT,
    summary="Delete a document set (only if not used by any trial)",
)
def delete_document_set(
    project_id: int,
    set_id: int,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    # 1. Permission check
    check_project_access(project_id, current_user, db, "write")

    # 2. Fetch the document set, ensure it belongs to project
    doc_set = db.execute(
        select(models.DocumentSet)
        .where(
            models.DocumentSet.id == set_id,
            models.DocumentSet.project_id == project_id,
        )
        .options(selectinload(models.DocumentSet.trials))
    ).scalar_one_or_none()

    if not doc_set:
        raise HTTPException(status_code=404, detail="Document set not found")

    # 3. Prevent deletion if any trial references this set
    if doc_set.trials and len(doc_set.trials) > 0:
        raise HTTPException(
            status_code=400,
            detail="Cannot delete document set: one or more trials reference it.",
        )

    # 4. Delete the set and its associations
    # Remove document associations (optional, for clean DB)
    db.execute(
        document_set_association.delete().where(
            document_set_association.c.document_set_id == set_id
        )
    )
    db.delete(doc_set)
    db.commit()
    return


@router.post(
    "/{project_id}/document-set/from-trial/{trial_id}",
    response_model=schemas.DocumentSet,
)
def create_document_set_from_trial(
    project_id: int,
    trial_id: int,
    set_data: schemas.DocumentSetFromTrial,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> schemas.DocumentSet:
    """Create a named document set from a trial's documents"""
    check_project_access(project_id, current_user, db, "write")

    # Get trial
    trial = db.execute(
        select(models.Trial).where(
            models.Trial.id == trial_id, models.Trial.project_id == project_id
        )
    ).scalar_one_or_none()

    if not trial:
        raise HTTPException(status_code=404, detail="Trial not found")

    # Create document set
    db_set = models.DocumentSet(
        project_id=project_id,
        name=set_data.name,
        description=set_data.description or f"Documents from trial #{trial_id}",
        tags=set_data.tags or [],
        is_auto_generated=False,
    )

    db.add(db_set)
    db.flush()

    # Add documents
    for doc_id in trial.document_ids:
        db.execute(
            document_set_association.insert().values(
                document_id=doc_id, document_set_id=db_set.id
            )
        )

    db.commit()
    db.refresh(db_set)

    return schemas.DocumentSet.model_validate(db_set)


# Updated document_set_endpoints.py - SQLite compatible version


@router.get(
    "/{project_id}/document-set/{set_id}/stats", response_model=schemas.DocumentSetStats
)
def get_document_set_stats(
    project_id: int,
    set_id: int,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> schemas.DocumentSetStats:
    """Get usage statistics for a document set"""
    check_project_access(project_id, current_user, db, "read")

    doc_set = db.execute(
        select(models.DocumentSet)
        .options(selectinload(models.DocumentSet.documents))
        .where(
            models.DocumentSet.id == set_id, models.DocumentSet.project_id == project_id
        )
    ).scalar_one_or_none()

    if not doc_set:
        raise HTTPException(status_code=404, detail="Document set not found")

    # Get document IDs in this set
    doc_ids = [doc.id for doc in doc_set.documents]

    if not doc_ids:
        return schemas.DocumentSetStats(
            trials_count=0, extractions_count=0, last_used=None
        )

    # For SQLite, we need to check JSON array membership differently
    # Count trials that contain ANY of these document IDs
    trials_with_docs = (
        db.execute(select(models.Trial).where(models.Trial.project_id == project_id))
        .scalars()
        .all()
    )

    # Filter trials that contain any of our document IDs
    trials_count = 0
    last_trial_date = None

    for trial in trials_with_docs:
        if trial.document_ids and any(
            doc_id in trial.document_ids for doc_id in doc_ids
        ):
            trials_count += 1
            if not last_trial_date or trial.created_at > last_trial_date:
                last_trial_date = trial.created_at

    # Count total extractions (trial results for these documents)
    extractions_count = (
        db.execute(
            select(func.count(models.TrialResult.id))
            .join(models.Trial)
            .where(
                models.Trial.project_id == project_id,
                models.TrialResult.document_id.in_(doc_ids),
            )
        ).scalar()
        or 0
    )

    return schemas.DocumentSetStats(
        trials_count=trials_count,
        extractions_count=extractions_count,
        last_used=last_trial_date,
    )


@router.post("/{project_id}/document-set/{set_id}/download-all")
def download_all_documents(
    project_id: int,
    set_id: int,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Download all documents in a set as a ZIP file"""
    check_project_access(project_id, current_user, db, "read")

    doc_set = db.execute(
        select(models.DocumentSet)
        .where(
            models.DocumentSet.id == set_id, models.DocumentSet.project_id == project_id
        )
        .options(
            selectinload(models.DocumentSet.documents).selectinload(
                models.Document.original_file
            )
        )
    ).scalar_one_or_none()

    if not doc_set:
        raise HTTPException(status_code=404, detail="Document set not found")

    # Create ZIP file in memory
    import io
    import zipfile

    zip_buffer = io.BytesIO()

    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
        for doc in doc_set.documents:
            if doc.original_file:
                try:
                    # Get file content
                    file_content = get_file(doc.original_file.file_uuid)

                    # Add to ZIP with original filename
                    zip_file.writestr(doc.original_file.file_name, file_content)
                except Exception as e:
                    # Log error but continue with other files
                    print(f"Error adding file {doc.original_file.file_name}: {e}")

    zip_buffer.seek(0)

    return StreamingResponse(
        zip_buffer,
        media_type="application/zip",
        headers={
            "Content-Disposition": f"attachment; filename={doc_set.name.replace(' ', '_')}_documents.zip"
        },
    )


@router.post("/{project_id}/schema", response_model=schemas.Schema)
def create_schema(
    project_id: int,
    schema: schemas.SchemaCreate,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> schemas.Schema:
    project: models.Project | None = db.execute(
        select(models.Project).where(models.Project.id == project_id)
    ).scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    if current_user.role != "admin" and project.owner_id != current_user.id:
        raise HTTPException(
            status_code=403, detail="Not authorized to create schemas for this project"
        )
    if not isinstance(schema.schema_definition, dict):
        raise HTTPException(status_code=400, detail="Invalid JSON schema")
    schema_db = models.Schema(**schema.model_dump(), project_id=project_id)
    db.add(schema_db)
    db.commit()
    db.refresh(schema_db)
    return schemas.Schema.model_validate(schema_db)


@router.get("/{project_id}/schema", response_model=list[schemas.Schema])
def get_schemas(
    project_id: int,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> list[schemas.Schema]:
    project: models.Project | None = db.execute(
        select(models.Project).where(models.Project.id == project_id)
    ).scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    if current_user.role != "admin" and project.owner_id != current_user.id:
        raise HTTPException(
            status_code=403, detail="Not authorized to access this project's schemas"
        )

    schemas_list = list(
        db.execute(select(models.Schema).where(models.Schema.project_id == project_id))
        .scalars()
        .all()
    )
    return [schemas.Schema.model_validate(schema) for schema in schemas_list]


@router.get("/{project_id}/schema/{schema_id}", response_model=schemas.Schema)
def get_schema(
    project_id: int,
    schema_id: int,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> schemas.Schema:
    project: models.Project | None = db.execute(
        select(models.Project).where(models.Project.id == project_id)
    ).scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    if current_user.role != "admin" and project.owner_id != current_user.id:
        raise HTTPException(
            status_code=403, detail="Not authorized to access this project's schemas"
        )

    schema: models.Schema | None = db.execute(
        select(models.Schema).where(
            models.Schema.project_id == project_id, models.Schema.id == schema_id
        )
    ).scalar_one_or_none()
    if not schema:
        raise HTTPException(status_code=404, detail="Schema not found")
    return schemas.Schema.model_validate(schema)


@router.put("/{project_id}/schema/{schema_id}", response_model=schemas.Schema)
def update_schema(
    project_id: int,
    schema_id: int,
    schema: schemas.SchemaUpdate,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> schemas.Schema:
    project: models.Project | None = db.execute(
        select(models.Project).where(models.Project.id == project_id)
    ).scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    if current_user.role != "admin" and project.owner_id != current_user.id:
        raise HTTPException(
            status_code=403, detail="Not authorized to update schemas for this project"
        )

    existing_schema: models.Schema | None = db.execute(
        select(models.Schema).where(
            models.Schema.project_id == project_id, models.Schema.id == schema_id
        )
    ).scalar_one_or_none()
    if not existing_schema:
        raise HTTPException(status_code=404, detail="Schema not found")

    if not isinstance(schema.schema_definition, dict):
        raise HTTPException(status_code=400, detail="Invalid JSON schema")

    for key, value in schema.model_dump(exclude_unset=True).items():
        setattr(existing_schema, key, value)

    db.add(existing_schema)
    db.commit()
    db.refresh(existing_schema)

    return schemas.Schema.model_validate(existing_schema)


@router.delete("/{project_id}/schema/{schema_id}", response_model=schemas.Schema)
def delete_schema(
    project_id: int,
    schema_id: int,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> schemas.Schema:
    project: models.Project | None = db.execute(
        select(models.Project).where(models.Project.id == project_id)
    ).scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    if current_user.role != "admin" and project.owner_id != current_user.id:
        raise HTTPException(
            status_code=403, detail="Not authorized to delete schemas from this project"
        )

    schema: models.Schema | None = db.execute(
        select(models.Schema).where(
            models.Schema.project_id == project_id, models.Schema.id == schema_id
        )
    ).scalar_one_or_none()
    if not schema:
        raise HTTPException(status_code=404, detail="Schema not found")

    # Check if the schema is referenced by any trials
    trials: list[models.Trial] = (
        db.execute(select(models.Trial).where(models.Trial.schema_id == schema_id))
        .scalars()
        .all()
    )

    if trials:
        raise HTTPException(
            status_code=400, detail="Cannot delete schema referenced by a trial"
        )

    db.delete(schema)
    db.commit()
    return schemas.Schema.model_validate(schema)


@router.post("/{project_id}/prompt", response_model=schemas.Prompt)
def create_prompt(
    project_id: int,
    prompt: schemas.PromptCreate,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> schemas.Prompt:
    project: models.Project | None = db.execute(
        select(models.Project).where(models.Project.id == project_id)
    ).scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    if current_user.role != "admin" and project.owner_id != current_user.id:
        raise HTTPException(
            status_code=403, detail="Not authorized to create prompts for this project"
        )

    # Validate prompt
    validate_prompt(prompt)

    prompt_db = models.Prompt(
        **prompt.model_dump(exclude={"project_id"}), project_id=project_id
    )
    db.add(prompt_db)
    db.commit()
    db.refresh(prompt_db)
    return schemas.Prompt.model_validate(prompt_db)


@router.get("/{project_id}/prompt", response_model=list[schemas.Prompt])
def get_prompts(
    project_id: int,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> list[schemas.Prompt]:
    project: models.Project | None = db.execute(
        select(models.Project).where(models.Project.id == project_id)
    ).scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    if current_user.role != "admin" and project.owner_id != current_user.id:
        raise HTTPException(
            status_code=403, detail="Not authorized to access this project's prompts"
        )

    prompts_list = list(
        db.execute(select(models.Prompt).where(models.Prompt.project_id == project_id))
        .scalars()
        .all()
    )
    return [schemas.Prompt.model_validate(prompt) for prompt in prompts_list]


@router.get("/{project_id}/prompt/{prompt_id}", response_model=schemas.Prompt)
def get_prompt(
    project_id: int,
    prompt_id: int,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> schemas.Prompt:
    project: models.Project | None = db.execute(
        select(models.Project).where(models.Project.id == project_id)
    ).scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    if current_user.role != "admin" and project.owner_id != current_user.id:
        raise HTTPException(
            status_code=403, detail="Not authorized to access this project's prompts"
        )

    prompt: models.Prompt | None = db.execute(
        select(models.Prompt).where(
            models.Prompt.project_id == project_id, models.Prompt.id == prompt_id
        )
    ).scalar_one_or_none()
    if not prompt:
        raise HTTPException(status_code=404, detail="Prompt not found")
    return schemas.Prompt.model_validate(prompt)


@router.put("/{project_id}/prompt/{prompt_id}", response_model=schemas.Prompt)
def update_prompt(
    project_id: int,
    prompt_id: int,
    prompt: schemas.PromptUpdate,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> schemas.Prompt:
    project: models.Project | None = db.execute(
        select(models.Project).where(models.Project.id == project_id)
    ).scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    if current_user.role != "admin" and project.owner_id != current_user.id:
        raise HTTPException(
            status_code=403, detail="Not authorized to update prompts for this project"
        )

    existing_prompt: models.Prompt | None = db.execute(
        select(models.Prompt).where(
            models.Prompt.project_id == project_id, models.Prompt.id == prompt_id
        )
    ).scalar_one_or_none()
    if not existing_prompt:
        raise HTTPException(status_code=404, detail="Prompt not found")

    # Create a temporary object to validate
    temp_prompt = schemas.PromptUpdate(
        system_prompt=prompt.system_prompt
        if prompt.system_prompt is not None
        else existing_prompt.system_prompt,
        user_prompt=prompt.user_prompt
        if prompt.user_prompt is not None
        else existing_prompt.user_prompt,
    )
    validate_prompt(temp_prompt)

    for key, value in prompt.model_dump(exclude_unset=True).items():
        setattr(existing_prompt, key, value)

    db.add(existing_prompt)
    db.commit()
    db.refresh(existing_prompt)

    return schemas.Prompt.model_validate(existing_prompt)


@router.delete("/{project_id}/prompt/{prompt_id}", response_model=schemas.Prompt)
def delete_prompt(
    project_id: int,
    prompt_id: int,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> schemas.Prompt:
    project: models.Project | None = db.execute(
        select(models.Project).where(models.Project.id == project_id)
    ).scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    if current_user.role != "admin" and project.owner_id != current_user.id:
        raise HTTPException(
            status_code=403, detail="Not authorized to delete prompts from this project"
        )

    prompt: models.Prompt | None = db.execute(
        select(models.Prompt).where(
            models.Prompt.project_id == project_id, models.Prompt.id == prompt_id
        )
    ).scalar_one_or_none()
    if not prompt:
        raise HTTPException(status_code=404, detail="Prompt not found")

    # Check if the prompt is referenced by any trials
    trial: models.Trial | None = db.execute(
        select(models.Trial).where(models.Trial.prompt_id == prompt_id)
    ).scalar_one_or_none()
    if trial:
        raise HTTPException(
            status_code=400, detail="Cannot delete prompt referenced by a trial"
        )

    db.delete(prompt)
    db.commit()
    return schemas.Prompt.model_validate(prompt)


@router.post("/{project_id}/trial", response_model=schemas.Trial)
def create_trial(
    project_id: int,
    trial: schemas.TrialCreate,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> schemas.Trial:
    # 1. Project, schema, prompt existence & authorization checks
    project: models.Project | None = db.execute(
        select(models.Project).where(models.Project.id == project_id)
    ).scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    if current_user.role != "admin" and project.owner_id != current_user.id:
        raise HTTPException(
            status_code=403, detail="Not authorized to create trials for this project"
        )

    schema: models.Schema | None = db.execute(
        select(models.Schema).where(models.Schema.id == trial.schema_id)
    ).scalar_one_or_none()
    if not schema:
        raise HTTPException(status_code=404, detail="Schema not found")

    prompt: models.Prompt | None = db.execute(
        select(models.Prompt).where(models.Prompt.id == trial.prompt_id)
    ).scalar_one_or_none()
    if not prompt:
        raise HTTPException(status_code=404, detail="Prompt not found")

    # 2. Decide document IDs to use: either direct list, or from document set
    document_ids: list[int] = []

    if trial.document_set_id is not None:
        # Document set mode
        document_set: models.DocumentSet | None = db.execute(
            select(models.DocumentSet).where(
                (models.DocumentSet.id == trial.document_set_id)
                & (models.DocumentSet.project_id == project_id)
            )
        ).scalar_one_or_none()
        if not document_set:
            raise HTTPException(
                status_code=404, detail="Document set not found in this project"
            )
        # Fetch all doc IDs from set
        document_ids = [doc.id for doc in document_set.documents]
        if not document_ids:
            raise HTTPException(status_code=400, detail="Document set is empty")
    elif trial.document_ids:
        # Explicit document IDs
        document_ids = trial.document_ids
        # Check all are in project
        existing_documents = (
            db.execute(
                select(models.Document.id).where(
                    models.Document.project_id == project_id
                )
            )
            .scalars()
            .all()
        )
        for document_id in document_ids:
            if document_id not in existing_documents:
                raise HTTPException(
                    status_code=404,
                    detail=f"Document with id {document_id} not found in project {project_id}",
                )
    else:
        raise HTTPException(
            status_code=400,
            detail="Either document_ids or document_set_id must be provided",
        )

    # 3. Use default config values if not set
    llm_model = trial.llm_model or settings.OPENAI_API_MODEL
    api_key = trial.api_key or settings.OPENAI_API_KEY
    base_url = trial.base_url or settings.OPENAI_API_BASE

    if llm_model is None or api_key is None or base_url is None:
        raise HTTPException(status_code=400, detail="LLM configuration is incomplete")

    # 4. Create trial object
    trial_db = models.Trial(
        name=trial.name,
        description=trial.description,
        schema_id=trial.schema_id,
        prompt_id=str(trial.prompt_id),
        project_id=project_id,
        llm_model=llm_model,
        api_key=api_key,
        base_url=base_url,
        document_ids=document_ids,
        document_set_id=trial.document_set_id
        if trial.document_set_id is not None
        else None,
        bypass_celery=trial.bypass_celery,
        advanced_options=trial.advanced_options or {},
    )

    # 5. Kick-off extraction ----------------------------------------------------
    trial_db.status = models.TrialStatus.PROCESSING
    trial_db.started_at = datetime.datetime.now(datetime.UTC)
    trial_db.progress = 0.0
    db.add(trial_db)
    db.commit()
    db.refresh(trial_db)

    if trial.bypass_celery:
        # synchronous (debug)
        from ....utils.info_extraction import (
            extract_info_single_doc,
            update_trial_progress,
        )

        for doc_id in document_ids:
            extract_info_single_doc(
                db_session=db,
                trial_id=trial_db.id,
                document_id=doc_id,
                llm_model=llm_model,
                api_key=api_key,
                base_url=base_url,
                schema_id=trial.schema_id,
                prompt_id=trial.prompt_id,
                project_id=project_id,
                advanced_options=trial_db.advanced_options,
            )
            update_trial_progress(db, trial_db.id)

        trial_db.status = models.TrialStatus.COMPLETED
        trial_db.finished_at = datetime.datetime.now(datetime.UTC)
        db.commit()
    else:
        from ....celery.info_extraction import extract_info_celery

        extract_info_celery.delay(
            trial_id=trial_db.id,
            document_ids=document_ids,
            llm_model=llm_model,
            api_key=api_key,
            base_url=base_url,
            schema_id=trial.schema_id,
            prompt_id=trial.prompt_id,
            project_id=project_id,
            advanced_options=trial_db.advanced_options,
        )

    return schemas.Trial.model_validate(trial_db)


@router.patch("/{project_id}/trial/{trial_id}", response_model=schemas.Trial)
def update_trial(
    project_id: int,
    trial_id: int,
    trial_update: schemas.TrialUpdate,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    # 1. Project existence and permission check
    project = db.execute(
        select(models.Project).where(models.Project.id == project_id)
    ).scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    if current_user.role != "admin" and project.owner_id != current_user.id:
        raise HTTPException(
            status_code=403, detail="Not authorized to update trials for this project"
        )

    # 2. Trial existence check
    trial = db.execute(
        select(models.Trial).where(
            models.Trial.project_id == project_id, models.Trial.id == trial_id
        )
    ).scalar_one_or_none()
    if not trial:
        raise HTTPException(status_code=404, detail="Trial not found")

    # 3. Update allowed fields
    updated = False
    if trial_update.name is not None:
        trial.name = trial_update.name
        updated = True
    if trial_update.description is not None:
        trial.description = trial_update.description
        updated = True

    if not updated:
        raise HTTPException(status_code=400, detail="No updatable fields provided")

    db.commit()
    db.refresh(trial)
    return schemas.Trial.model_validate(trial)


@router.post("/{project_id}/trial/{trial_id}/cancel", response_model=schemas.Trial)
def cancel_trial(
    project_id: int,
    trial_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    check_project_access(project_id, current_user, db)
    trial = db.get(models.Trial, trial_id)
    if not trial or trial.project_id != project_id:
        raise HTTPException(status_code=404, detail="Trial not found")
    if trial.is_cancelled or trial.status in ("completed", "failed", "cancelled"):
        raise HTTPException(status_code=400, detail="Trial cannot be cancelled")

    trial.is_cancelled = True
    trial.status = models.TrialStatus.CANCELLED
    db.commit()
    db.refresh(trial)
    return schemas.Trial.model_validate(trial)


@router.delete("/{project_id}/trial/{trial_id}", response_model=schemas.Trial)
def delete_trial(
    project_id: int,
    trial_id: int,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> schemas.Trial:
    # Project and permission checks
    project: models.Project | None = db.execute(
        select(models.Project).where(models.Project.id == project_id)
    ).scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    if current_user.role != "admin" and project.owner_id != current_user.id:
        raise HTTPException(
            status_code=403, detail="Not authorized to delete trials for this project"
        )

    # Fetch the trial
    trial: models.Trial | None = db.execute(
        select(models.Trial).where(
            models.Trial.project_id == project_id, models.Trial.id == trial_id
        )
    ).scalar_one_or_none()
    if not trial:
        raise HTTPException(status_code=404, detail="Trial not found")

    # --- SERIALIZE before delete ---
    trial_data = schemas.Trial.model_validate(trial)

    try:
        # Delete all evaluations (and their metrics, via cascade)
        evaluations = (
            db.execute(
                select(models.Evaluation).where(models.Evaluation.trial_id == trial_id)
            )
            .scalars()
            .all()
        )
        for evaluation in evaluations:
            db.delete(evaluation)

        # Delete all trial results for this trial
        results = (
            db.execute(
                select(models.TrialResult).where(
                    models.TrialResult.trial_id == trial_id
                )
            )
            .scalars()
            .all()
        )
        for result in results:
            db.delete(result)

        # Delete the trial itself
        db.delete(trial)
        db.commit()
    except SQLAlchemyError as e:
        db.rollback()
        raise HTTPException(
            status_code=500, detail=f"Database error during deletion: {e}"
        )

    return trial_data


@router.get("/{project_id}/trial", response_model=list[schemas.Trial])
def get_trials(
    project_id: int,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> list[schemas.Trial]:
    project: models.Project | None = db.execute(
        select(models.Project).where(models.Project.id == project_id)
    ).scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    if current_user.role != "admin" and project.owner_id != current_user.id:
        raise HTTPException(
            status_code=403, detail="Not authorized to access this project's trials"
        )

    trials = list(
        db.execute(select(models.Trial).where(models.Trial.project_id == project_id))
        .scalars()
        .all()
    )
    return [schemas.Trial.model_validate(trial) for trial in trials]


@router.get("/{project_id}/trial/{trial_id}", response_model=schemas.Trial)
def get_trial(
    project_id: int,
    trial_id: int,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> schemas.Trial:
    project: models.Project | None = db.execute(
        select(models.Project).where(models.Project.id == project_id)
    ).scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    if current_user.role != "admin" and project.owner_id != current_user.id:
        raise HTTPException(
            status_code=403, detail="Not authorized to access this project's trials"
        )

    trial: models.Trial | None = db.execute(
        select(models.Trial).where(
            models.Trial.project_id == project_id, models.Trial.id == trial_id
        )
    ).scalar_one_or_none()

    if trial:
        trial.results = cast(
            list[models.TrialResult],
            (
                db.execute(
                    select(models.TrialResult).where(
                        models.TrialResult.trial_id == trial_id
                    )
                )
                .scalars()
                .all()
            ),
        )

    if not trial:
        raise HTTPException(status_code=404, detail="Trial not found")
    return schemas.Trial.model_validate(trial)


@router.get("/{project_id}/trial/{trial_id}/download", response_class=Response)
def download_trial_results(
    *,
    db: Session = Depends(get_db),
    project_id: int,
    trial_id: int,
    format: str = Query("json", enum=["json", "csv"]),
    include_content: bool = Query(True),
    current_user: "models.User" = Depends(get_current_user),
) -> Response:
    """
    Download trial results, with a separate metadata.json for trial/prompt/schema metadata.
    """

    def filter_sensitive_keys(d, blacklist=("api_key",)):
        if not d:
            return {}
        return {k: v for k, v in d.items() if k not in blacklist}

    # --- Permissions ---
    project = db.execute(
        select(models.Project).where(models.Project.id == project_id)
    ).scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    if current_user.role != "admin" and project.owner_id != current_user.id:
        raise HTTPException(
            status_code=403, detail="Not authorized to access this project's trials"
        )
    trial = db.execute(
        select(models.Trial).where(
            models.Trial.project_id == project_id, models.Trial.id == trial_id
        )
    ).scalar_one_or_none()
    if not trial:
        raise HTTPException(status_code=404, detail="Trial not found")

    # --- Related: prompt, schema ---
    prompt = db.execute(
        select(models.Prompt).where(models.Prompt.id == trial.prompt_id)
    ).scalar_one_or_none()
    schema = db.execute(
        select(models.Schema).where(models.Schema.id == trial.schema_id)
    ).scalar_one_or_none()

    results = list(
        db.execute(
            select(models.TrialResult).where(models.TrialResult.trial_id == trial_id)
        )
        .scalars()
        .all()
    )
    if not results:
        raise HTTPException(status_code=404, detail="No results found for this trial")

    document_cache = {}
    file_cache = {}
    preprocessing_config_cache = {}
    all_meta_keys = set()
    all_result_keys = set()
    all_prep_keys = set()
    # all_trial_keys = set()
    # all_prompt_keys = set()
    # all_schema_keys = set()

    for result in results:
        doc = db.execute(
            select(models.Document).where(models.Document.id == result.document_id)
        ).scalar_one_or_none()
        document_cache[result.document_id] = doc
        if doc:
            if doc.original_file_id and doc.original_file_id not in file_cache:
                file_cache[doc.original_file_id] = db.execute(
                    select(models.File).where(models.File.id == doc.original_file_id)
                ).scalar_one_or_none()
            prep_conf = None
            if doc.preprocessing_config_id:
                prep_conf = db.execute(
                    select(models.PreprocessingConfiguration).where(
                        models.PreprocessingConfiguration.id
                        == doc.preprocessing_config_id
                    )
                ).scalar_one_or_none()
                if prep_conf:
                    preprocessing_config_cache[doc.preprocessing_config_id] = prep_conf
                    all_prep_keys.update(
                        _extract_keys(filter_sensitive_keys(prep_conf.__dict__))
                    )
            all_meta_keys.update(_extract_keys(doc.meta_data or {}))
            all_result_keys.update(_extract_keys(result.result or {}))

    # Prepare metadata for metadata.json (top-level only, filter sensitive)
    trial_dict = filter_sensitive_keys(
        {
            k: v
            for k, v in trial.__dict__.items()
            if not k.startswith("_")
            and isinstance(v, (str, int, float, bool, dict, list, type(None)))
        }
    )
    prompt_dict = filter_sensitive_keys(
        {
            k: v
            for k, v in (prompt.__dict__ if prompt else {}).items()
            if not k.startswith("_")
            and isinstance(v, (str, int, float, bool, dict, list, type(None)))
        }
    )
    schema_dict = filter_sensitive_keys(
        {
            k: v
            for k, v in (schema.__dict__ if schema else {}).items()
            if not k.startswith("_")
            and isinstance(v, (str, int, float, bool, dict, list, type(None)))
        }
    )

    # Remove fields that don't serialize well (e.g. relationships)
    for d in (trial_dict, prompt_dict, schema_dict):
        for key in list(d.keys()):
            if isinstance(d[key], (dict, list)):
                continue
            try:
                json.dumps(d[key])
            except Exception:
                d.pop(key)

    # --- JSON Format: Each Document as JSON in ZIP, metadata.json separate ---
    if format == "json":
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zipf:
            # Write metadata.json first
            metadata_json = {
                "trial": trial_dict,
                "prompt": prompt_dict,
                "schema": schema_dict,
            }
            zipf.writestr(
                "metadata.json", json.dumps(metadata_json, indent=2, ensure_ascii=False)
            )

            added_files = set()
            for result in results:
                document = document_cache.get(result.document_id)
                if not document:
                    continue

                file = None
                document_name = document.document_name
                if document.original_file_id:
                    file = file_cache.get(document.original_file_id)
                    if not document_name and file:
                        document_name = file.file_name

                prep_conf = None
                if document.preprocessing_config_id:
                    prep_obj = preprocessing_config_cache.get(
                        document.preprocessing_config_id
                    )
                    if prep_obj:
                        prep_conf = filter_sensitive_keys(
                            {
                                k: v
                                for k, v in prep_obj.__dict__.items()
                                if not k.startswith("_")
                                and isinstance(
                                    v, (str, int, float, bool, dict, list, type(None))
                                )
                            }
                        )

                result_data = {
                    "result": result.result,
                    "document_id": result.document_id,
                    "document_name": document_name,
                    "file_name": file.file_name if file else None,
                    "created_at": result.created_at.isoformat(),
                    "document_metadata": document.meta_data or {},
                    "preprocessing": prep_conf or {},
                }
                if include_content:
                    result_data["content"] = document.text
                    file_id = document.preprocessed_file_id or document.original_file_id
                    if file_id:
                        file_to_add = db.execute(
                            select(models.File).where(models.File.id == file_id)
                        ).scalar_one_or_none()
                        if file_to_add and file_id not in added_files:
                            added_files.add(file_id)
                            file_content = get_file(file_to_add.file_uuid)
                            file_path = (
                                f"files/{file_to_add.file_uuid}_{file_to_add.file_name}"
                            )
                            zipf.writestr(file_path, file_content)

                file_base = document_name or f"document_{result.document_id}"
                safe_base = "".join(
                    c for c in file_base if c.isalnum() or c in " ._-"
                ).rstrip()
                if not safe_base:
                    safe_base = f"document_{result.document_id}"
                json_filename = f"{safe_base}.json"
                zipf.writestr(
                    json_filename,
                    json.dumps(result_data, indent=2, ensure_ascii=False),
                )
        zip_buffer.seek(0)
        return Response(
            content=zip_buffer.getvalue(),
            media_type="application/zip",
            headers={
                "Content-Disposition": f"attachment; filename=trial_{trial_id}_results.zip"
            },
        )

    # --- CSV Format ---
    elif format == "csv":
        meta_keys = sorted(all_meta_keys)
        prep_keys = sorted(all_prep_keys)
        result_keys = sorted(all_result_keys)
        trial_keys = sorted(_extract_keys(trial_dict))
        prompt_keys = sorted(_extract_keys(prompt_dict))
        schema_keys = sorted(_extract_keys(schema_dict))

        def meta_flatten(md):
            return flatten_dict(md or {})

        def result_flatten(res):
            return flatten_dict(res or {})

        def prep_flatten(pc):
            return flatten_dict(pc or {})

        def flatten_one(d):
            return flatten_dict(d or {})

        if include_content:
            output = io.BytesIO()
            with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as zipf:
                header = (
                    [
                        "document_id",
                        "document_name",
                        "file_name",
                        "created_at",
                        "document_content",
                    ]
                    + [f"meta.{k}" for k in meta_keys]
                    + [f"preprocessing.{k}" for k in prep_keys]
                    + [f"trial.{k}" for k in trial_keys]
                    + [f"prompt.{k}" for k in prompt_keys]
                    + [f"schema.{k}" for k in schema_keys]
                    + [f"result.{k}" for k in result_keys]
                )
                csv_output = io.StringIO()
                writer = csv.DictWriter(csv_output, fieldnames=header)
                writer.writeheader()
                added_files = set()
                for result in results:
                    row = {
                        "document_id": result.document_id,
                        "document_name": "",
                        "file_name": "",
                        "created_at": result.created_at.isoformat(),
                        "document_content": "",
                    }
                    document = document_cache.get(result.document_id)
                    file = None
                    document_name = ""
                    file_name = ""
                    prep_conf = {}
                    if document:
                        document_name = document.document_name
                        if document.original_file_id:
                            file = file_cache.get(document.original_file_id)
                            if not document_name and file:
                                document_name = file.file_name
                        file_name = file.file_name if file else ""
                        row["document_content"] = document.text or ""
                        if document.preprocessing_config_id:
                            prep_obj = preprocessing_config_cache.get(
                                document.preprocessing_config_id
                            )
                            if prep_obj:
                                prep_conf = filter_sensitive_keys(
                                    {
                                        k: v
                                        for k, v in prep_obj.__dict__.items()
                                        if not k.startswith("_")
                                        and isinstance(
                                            v,
                                            (
                                                str,
                                                int,
                                                float,
                                                bool,
                                                dict,
                                                list,
                                                type(None),
                                            ),
                                        )
                                    }
                                )
                        file_id = (
                            document.preprocessed_file_id or document.original_file_id
                        )
                        if file_id:
                            file_to_add = db.execute(
                                select(models.File).where(models.File.id == file_id)
                            ).scalar_one_or_none()
                            if file_to_add and file_id not in added_files:
                                added_files.add(file_id)
                                file_content = get_file(file_to_add.file_uuid)
                                file_path = f"files/{file_to_add.file_uuid}_{file_to_add.file_name}"
                                zipf.writestr(file_path, file_content)
                    row["document_name"] = document_name or ""
                    row["file_name"] = file_name or ""
                    # Add metadata
                    meta_flat = meta_flatten(document.meta_data if document else {})
                    for k in meta_keys:
                        row[f"meta.{k}"] = meta_flat.get(k, "")
                    prep_flat = prep_flatten(prep_conf)
                    for k in prep_keys:
                        row[f"preprocessing.{k}"] = prep_flat.get(k, "")
                    # Add trial, prompt, schema (same for all)
                    trial_flat = flatten_one(trial_dict)
                    for k in trial_keys:
                        row[f"trial.{k}"] = trial_flat.get(k, "")
                    prompt_flat = flatten_one(prompt_dict)
                    for k in prompt_keys:
                        row[f"prompt.{k}"] = prompt_flat.get(k, "")
                    schema_flat = flatten_one(schema_dict)
                    for k in schema_keys:
                        row[f"schema.{k}"] = schema_flat.get(k, "")
                    res_flat = result_flatten(result.result)
                    for k in result_keys:
                        row[f"result.{k}"] = res_flat.get(k, "")
                    writer.writerow(row)
                zipf.writestr("results.csv", csv_output.getvalue())
            output.seek(0)
            return Response(
                content=output.getvalue(),
                media_type="application/zip",
                headers={
                    "Content-Disposition": f"attachment; filename=trial_{trial_id}_results.zip"
                },
            )
        else:
            output = io.StringIO()
            header = (
                ["document_id", "document_name", "file_name", "created_at"]
                + [f"meta.{k}" for k in meta_keys]
                + [f"preprocessing.{k}" for k in prep_keys]
                + [f"trial.{k}" for k in trial_keys]
                + [f"prompt.{k}" for k in prompt_keys]
                + [f"schema.{k}" for k in schema_keys]
                + [f"result.{k}" for k in result_keys]
            )
            writer = csv.DictWriter(output, fieldnames=header)
            writer.writeheader()
            for result in results:
                row = {
                    "document_id": result.document_id,
                    "document_name": "",
                    "file_name": "",
                    "created_at": result.created_at.isoformat(),
                }
                document = document_cache.get(result.document_id)
                file = None
                document_name = ""
                file_name = ""
                prep_conf = {}
                if document:
                    document_name = document.document_name
                    if document.original_file_id:
                        file = file_cache.get(document.original_file_id)
                        if not document_name and file:
                            document_name = file.file_name
                    file_name = file.file_name if file else ""
                    if document.preprocessing_config_id:
                        prep_obj = preprocessing_config_cache.get(
                            document.preprocessing_config_id
                        )
                        if prep_obj:
                            prep_conf = filter_sensitive_keys(
                                {
                                    k: v
                                    for k, v in prep_obj.__dict__.items()
                                    if not k.startswith("_")
                                    and isinstance(
                                        v,
                                        (str, int, float, bool, dict, list, type(None)),
                                    )
                                }
                            )
                row["document_name"] = document_name or ""
                row["file_name"] = file_name or ""
                meta_flat = flatten_dict(document.meta_data if document else {})
                for k in meta_keys:
                    row[f"meta.{k}"] = meta_flat.get(k, "")
                prep_flat = flatten_dict(prep_conf)
                for k in prep_keys:
                    row[f"preprocessing.{k}"] = prep_flat.get(k, "")
                trial_flat = flatten_dict(trial_dict)
                for k in trial_keys:
                    row[f"trial.{k}"] = trial_flat.get(k, "")
                prompt_flat = flatten_dict(prompt_dict)
                for k in prompt_keys:
                    row[f"prompt.{k}"] = prompt_flat.get(k, "")
                schema_flat = flatten_dict(schema_dict)
                for k in schema_keys:
                    row[f"schema.{k}"] = schema_flat.get(k, "")
                res_flat = flatten_dict(result.result)
                for k in result_keys:
                    row[f"result.{k}"] = res_flat.get(k, "")
                writer.writerow(row)
            return Response(
                content=output.getvalue(),
                media_type="text/csv",
                headers={
                    "Content-Disposition": f"attachment; filename=trial_{trial_id}_results.csv"
                },
            )

    raise HTTPException(status_code=404, detail="No results found for this trial")


def _extract_keys(d, parent_key="", sep="_"):
    """Extract all keys from a nested dictionary."""
    keys = []
    for k, v in d.items():
        new_key = f"{parent_key}{sep}{k}" if parent_key else k
        if isinstance(v, dict):
            keys.extend(_extract_keys(v, new_key, sep=sep))
        else:
            keys.append(new_key)
    return keys


@router.get("/llm/models", response_model=dict[str, Any])
def get_available_llm_models(
    api_key: str | None = settings.OPENAI_API_KEY,
    base_url: str | None = settings.OPENAI_API_BASE,
    current_user: models.User = Depends(get_current_user),
) -> dict[str, Any]:
    if api_key is None or base_url is None:
        return {
            "success": False,
            "models": [],
            "message": "LLM configuration is incomplete",
            "error_type": "incomplete_config",
        }
    return get_available_models(api_key, base_url)


@router.post("/llm/test-connection", response_model=dict[str, Any])
def test_api_connection_endpoint(
    api_key: str | None = settings.OPENAI_API_KEY,
    base_url: str | None = settings.OPENAI_API_BASE,
    current_user: models.User = Depends(get_current_user),
) -> dict[str, Any]:
    if api_key is None or base_url is None:
        return {
            "success": False,
            "message": "LLM configuration is incomplete. Please provide API key and base URL.",
            "error_type": "incomplete_config",
        }
    return test_api_connection(api_key, base_url)


@router.post("/llm/test-model", response_model=dict[str, Any])
def test_llm_model_endpoint(
    api_key: str | None = settings.OPENAI_API_KEY,
    base_url: str | None = settings.OPENAI_API_BASE,
    llm_model: str | None = settings.OPENAI_API_MODEL,
    current_user: models.User = Depends(get_current_user),
) -> dict[str, Any]:
    if api_key is None or base_url is None or llm_model is None:
        return {
            "success": False,
            "message": "LLM configuration is incomplete. Please provide API key, base URL, and model.",
            "error_type": "incomplete_config",
        }
    return test_llm_connection(api_key, base_url, llm_model)


# Add this new endpoint to your router:


@router.post("/llm/test-model-schema", response_model=dict[str, Any])
def test_model_with_schema_endpoint(
    api_key: str | None = settings.OPENAI_API_KEY,
    base_url: str | None = settings.OPENAI_API_BASE,
    llm_model: str | None = settings.OPENAI_API_MODEL,
    schema_id: int | None = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
) -> dict[str, Any]:
    """Test if a model supports structured output with a specific schema"""
    if api_key is None or base_url is None or llm_model is None:
        return {
            "success": False,
            "message": "LLM configuration is incomplete. Please provide API key, base URL, and model.",
            "error_type": "incomplete_config",
        }

    if schema_id is None:
        return {
            "success": False,
            "message": "Schema ID is required for testing structured output.",
            "error_type": "missing_schema",
        }

    # Get the schema
    schema = db.execute(
        select(models.Schema).where(models.Schema.id == schema_id)
    ).scalar_one_or_none()

    if not schema:
        return {
            "success": False,
            "message": "Schema not found.",
            "error_type": "schema_not_found",
        }

    # Test the model with structured output
    return test_model_with_schema(
        api_key, base_url, llm_model, schema.schema_definition
    )


@router.get("/llm/test-vlm-image-support")
def test_vlm_image_support(
    *,
    db: Session = Depends(get_db),
    model: str,
    api_key: str | None = None,
    base_url: str | None = None,
    current_user: models.User = Depends(get_current_user),
) -> dict:
    """
    Test if a VLM model supports image input.
    All params are optional (except model).
    """

    # Use default settings if not provided
    api_key = api_key or settings.OPENAI_API_KEY
    base_url = base_url or settings.OPENAI_API_BASE

    if not api_key or not base_url or not model:
        return {
            "supported": False,
            "message": "Configuration incomplete: api_key, base_url, and model are required",
        }

    try:
        from ....utils.helpers import test_remote_image_support

        api_url = base_url
        if not api_url.endswith("/chat/completions"):
            api_url = api_url.rstrip("/") + "/chat/completions"

        supported = test_remote_image_support(
            api_url=api_url, model=model, api_key=api_key
        )

        return {
            "supported": supported,
            "message": "Model supports image input"
            if supported
            else "Model does not support image input",
        }
    except Exception as e:
        return {"supported": False, "message": f"Test failed: {str(e)}"}


@router.post("/{project_id}/groundtruth", response_model=schemas.GroundTruth)
async def upload_groundtruth(
    *,
    db: Session = Depends(get_db),
    project_id: int,
    files: List[UploadFile] = File(None),  # For multiple files
    file: UploadFile = File(None),  # For single file (backward compatibility)
    name: str = Form(None),
    format: str = Form(...),
    multiple_json: bool = Form(False),
    current_user: models.User = Depends(get_current_user),
) -> schemas.GroundTruth:
    """Upload ground truth file for evaluation."""
    project: models.Project | None = db.execute(
        select(models.Project).where(models.Project.id == project_id)
    ).scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    if current_user.role != "admin" and project.owner_id != current_user.id:
        raise HTTPException(
            status_code=403,
            detail="Not authorized to upload ground truth to this project",
        )
    # Validate format
    if format not in ["json", "csv", "xlsx", "zip"]:
        raise HTTPException(
            status_code=400, detail="Invalid format. Must be json, csv, xlsx, or zip"
        )
    # Save the file
    if multiple_json and files:
        # Create a ZIP file in memory containing all JSON files
        import io
        import zipfile

        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, "w") as zf:
            for idx, json_file in enumerate(files):
                content = await json_file.read()
                # Validate it's valid JSON
                try:
                    json.loads(content)
                except json.JSONDecodeError:
                    raise HTTPException(
                        status_code=400,
                        detail=f"File {json_file.filename} is not valid JSON",
                    )
                zf.writestr(json_file.filename, content)

        zip_buffer.seek(0)
        file_content = zip_buffer.read()
        format = "zip"  # Treat as ZIP internally
    else:
        # Single file upload (existing logic)
        upload_file = file or (files[0] if files else None)
        if not upload_file:
            raise HTTPException(status_code=400, detail="No file provided")
        file_content = await upload_file.read()
    file_uuid = save_file(file_content)
    # Create ground truth record
    gt = models.GroundTruth(
        project_id=project_id,
        name=name or file.filename,
        format=format,
        file_uuid=file_uuid,
    )
    db.add(gt)
    db.commit()
    db.refresh(gt)
    return schemas.GroundTruth.model_validate(gt)


@router.get(
    "/{project_id}/groundtruth/{groundtruth_id}", response_model=schemas.GroundTruth
)
def get_groundtruth(
    *,
    db: Session = Depends(get_db),
    project_id: int,
    groundtruth_id: int,
    current_user: models.User = Depends(get_current_user),
) -> schemas.GroundTruth:
    """Get a specific ground truth file."""
    project: models.Project | None = db.execute(
        select(models.Project).where(models.Project.id == project_id)
    ).scalar_one_or_none()

    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    if current_user.role != "admin" and project.owner_id != current_user.id:
        raise HTTPException(
            status_code=403,
            detail="Not authorized to access this project's ground truth",
        )

    groundtruth: models.GroundTruth | None = db.execute(
        select(models.GroundTruth).where(
            models.GroundTruth.project_id == project_id,
            models.GroundTruth.id == groundtruth_id,
        )
    ).scalar_one_or_none()

    if not groundtruth:
        raise HTTPException(status_code=404, detail="Ground truth not found")

    return schemas.GroundTruth.model_validate(groundtruth)


@router.delete(
    "/{project_id}/groundtruth/{groundtruth_id}", response_model=schemas.GroundTruth
)
def delete_groundtruth(
    *,
    db: Session = Depends(get_db),
    project_id: int,
    groundtruth_id: int,
    current_user: models.User = Depends(get_current_user),
) -> schemas.GroundTruth:
    """Delete a ground truth file."""
    project: models.Project | None = db.execute(
        select(models.Project).where(models.Project.id == project_id)
    ).scalar_one_or_none()

    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    if current_user.role != "admin" and project.owner_id != current_user.id:
        raise HTTPException(
            status_code=403,
            detail="Not authorized to delete this project's ground truth",
        )

    groundtruth: models.GroundTruth | None = db.execute(
        select(models.GroundTruth).where(
            models.GroundTruth.project_id == project_id,
            models.GroundTruth.id == groundtruth_id,
        )
    ).scalar_one_or_none()

    if not groundtruth:
        raise HTTPException(status_code=404, detail="Ground truth not found")

    # Delete the file content from storage
    try:
        remove_file(groundtruth.file_uuid)
    except FileNotFoundError:
        # Continue deletion even if file not found in storage
        pass

    # Remove any evaluations using this ground truth
    evaluations = (
        db.execute(
            select(models.Evaluation).where(
                models.Evaluation.groundtruth_id == groundtruth_id
            )
        )
        .scalars()
        .all()
    )

    for evaluation in evaluations:
        db.delete(evaluation)

    db.delete(groundtruth)
    db.commit()

    return schemas.GroundTruth.model_validate(groundtruth)


@router.get("/{project_id}/groundtruth", response_model=list[schemas.GroundTruth])
def get_groundtruth_files(
    *,
    db: Session = Depends(get_db),
    project_id: int,
    current_user: models.User = Depends(get_current_user),
) -> list[schemas.GroundTruth]:
    """Get all ground truth files for a project."""
    project: models.Project | None = db.execute(
        select(models.Project).where(models.Project.id == project_id)
    ).scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    if current_user.role != "admin" and project.owner_id != current_user.id:
        raise HTTPException(
            status_code=403,
            detail="Not authorized to access this project's ground truth files",
        )
    ground_truths = list(
        db.execute(
            select(models.GroundTruth).where(
                models.GroundTruth.project_id == project_id
            )
        )
        .scalars()
        .all()
    )
    return [schemas.GroundTruth.model_validate(gt) for gt in ground_truths]


@router.put(
    "/{project_id}/groundtruth/{groundtruth_id}", response_model=schemas.GroundTruth
)
def update_groundtruth(
    *,
    db: Session = Depends(get_db),
    project_id: int,
    groundtruth_id: int,
    file: UploadFile = File(None),
    name: str = Form(None),
    comparison_options: str = Form(None),
    current_user: models.User = Depends(get_current_user),
) -> schemas.GroundTruth:
    project: models.Project | None = db.execute(
        select(models.Project).where(models.Project.id == project_id)
    ).scalar_one_or_none()

    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    if current_user.role != "admin" and project.owner_id != current_user.id:
        raise HTTPException(
            status_code=403,
            detail="Not authorized to update this project's ground truth",
        )

    groundtruth: models.GroundTruth | None = db.execute(
        select(models.GroundTruth).where(
            models.GroundTruth.project_id == project_id,
            models.GroundTruth.id == groundtruth_id,
        )
    ).scalar_one_or_none()

    if not groundtruth:
        raise HTTPException(status_code=404, detail="Ground truth not found")

    if name:
        groundtruth.name = name

    if comparison_options:
        try:
            groundtruth.comparison_options = json.loads(comparison_options)
        except json.JSONDecodeError:
            raise HTTPException(
                status_code=400, detail="Invalid comparison options JSON"
            )

    if file:
        try:
            remove_file(groundtruth.file_uuid)
        except FileNotFoundError:
            pass

        file_content = file.file.read()
        file_uuid = save_file(file_content)
        groundtruth.file_uuid = file_uuid

        evaluations = (
            db.execute(
                select(models.Evaluation).where(
                    models.Evaluation.groundtruth_id == groundtruth_id
                )
            )
            .scalars()
            .all()
        )
        for evaluation in evaluations:
            db.delete(evaluation)

    groundtruth.updated_at = func.now()
    db.add(groundtruth)
    db.commit()
    db.refresh(groundtruth)

    return schemas.GroundTruth.model_validate(groundtruth)


# Add to the ground truth endpoints
@router.put(
    "/{project_id}/groundtruth/{groundtruth_id}/id-column",
    response_model=schemas.GroundTruth,
)
def update_ground_truth_id_column(
    *,
    db: Session = Depends(get_db),
    project_id: int,
    groundtruth_id: int,
    id_column: str = Body(..., embed=True),
    current_user: models.User = Depends(get_current_user),
) -> schemas.GroundTruth:
    """Update the ID column for CSV/XLSX ground truth files."""
    # Validate project access
    project: models.Project | None = db.execute(
        select(models.Project).where(models.Project.id == project_id)
    ).scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    if current_user.role != "admin" and project.owner_id != current_user.id:
        raise HTTPException(
            status_code=403, detail="Not authorized to update ground truth"
        )

    # Get ground truth
    groundtruth: models.GroundTruth | None = db.execute(
        select(models.GroundTruth).where(
            models.GroundTruth.project_id == project_id,
            models.GroundTruth.id == groundtruth_id,
        )
    ).scalar_one_or_none()
    if not groundtruth:
        raise HTTPException(status_code=404, detail="Ground truth not found")

    # Only allow for CSV/XLSX formats
    if groundtruth.format not in ["csv", "xlsx"]:
        raise HTTPException(
            status_code=400,
            detail=f"ID column mapping is only supported for CSV and XLSX formats, not {groundtruth.format}",
        )

    # Update ID column
    groundtruth.id_column_name = id_column

    # Clear data cache to force re-parsing with new ID column
    groundtruth.data_cache = None

    # Invalidate related evaluations
    db.execute(
        delete(models.Evaluation).where(
            models.Evaluation.groundtruth_id == groundtruth_id
        )
    )

    db.commit()
    db.refresh(groundtruth)

    return schemas.GroundTruth.model_validate(groundtruth)


@router.post(
    "/{project_id}/groundtruth/{groundtruth_id}/schema/{schema_id}/mapping",
    response_model=schemas.GroundTruth,
)
def configure_field_mapping(
    *,
    db: Session = Depends(get_db),
    project_id: int,
    groundtruth_id: int,
    schema_id: int,
    mappings: list[schemas.FieldMappingCreate] = Body(...),
    current_user: models.User = Depends(get_current_user),
) -> schemas.GroundTruth:
    """Configure field mappings for a specific groundtruth-schema combination."""
    # Validate project access
    project: models.Project | None = db.execute(
        select(models.Project).where(models.Project.id == project_id)
    ).scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    if current_user.role != "admin" and project.owner_id != current_user.id:
        raise HTTPException(
            status_code=403, detail="Not authorized to configure field mappings"
        )

    # Validate groundtruth exists
    groundtruth: models.GroundTruth | None = db.execute(
        select(models.GroundTruth).where(
            models.GroundTruth.project_id == project_id,
            models.GroundTruth.id == groundtruth_id,
        )
    ).scalar_one_or_none()
    if not groundtruth:
        raise HTTPException(status_code=404, detail="Ground truth not found")

    # Validate schema exists
    schema: models.Schema | None = db.execute(
        select(models.Schema).where(
            models.Schema.project_id == project_id,
            models.Schema.id == schema_id,
        )
    ).scalar_one_or_none()
    if not schema:
        raise HTTPException(status_code=404, detail="Schema not found")

    # Delete existing mappings for this groundtruth-schema combination
    db.execute(
        delete(models.FieldMapping).where(
            models.FieldMapping.ground_truth_id == groundtruth_id,
            models.FieldMapping.schema_id == schema_id,
        )
    )

    # Create new mappings
    for mapping_data in mappings:
        mapping = models.FieldMapping(
            ground_truth_id=groundtruth_id, **mapping_data.model_dump()
        )
        db.add(mapping)

    # Invalidate evaluations for this specific groundtruth-schema combination
    db.execute(
        delete(models.Evaluation).where(
            models.Evaluation.groundtruth_id == groundtruth_id,
            models.Evaluation.trial_id.in_(
                select(models.Trial.id).where(models.Trial.schema_id == schema_id)
            ),
        )
    )

    db.commit()
    db.refresh(groundtruth)
    return schemas.GroundTruth.model_validate(groundtruth)


@router.get("/{project_id}/schema/{schema_id}/field_types", response_model=dict)
def get_schema_field_types(
    *,
    db: Session = Depends(get_db),
    project_id: int,
    schema_id: int,
    current_user: models.User = Depends(get_current_user),
) -> dict:
    project: models.Project | None = db.execute(
        select(models.Project).where(models.Project.id == project_id)
    ).scalar_one_or_none()

    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    if current_user.role != "admin" and project.owner_id != current_user.id:
        raise HTTPException(
            status_code=403, detail="Not authorized to access this project's schemas"
        )

    schema: models.Schema | None = db.execute(
        select(models.Schema).where(
            models.Schema.project_id == project_id, models.Schema.id == schema_id
        )
    ).scalar_one_or_none()

    if not schema:
        raise HTTPException(status_code=404, detail="Schema not found")

    field_types = {}
    extract_field_types_from_schema(schema.schema_definition, field_types)

    return field_types


@router.get("/{project_id}/evaluation", response_model=list[schemas.Evaluation])
def get_evaluations(
    *,
    db: Session = Depends(get_db),
    project_id: int,
    groundtruth_id: int = Query(...),
    current_user: models.User = Depends(get_current_user),
) -> list[schemas.Evaluation]:
    """Get evaluation results for all trials using a specific ground truth."""
    project: models.Project | None = db.execute(
        select(models.Project).where(models.Project.id == project_id)
    ).scalar_one_or_none()

    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    if current_user.role != "admin" and project.owner_id != current_user.id:
        raise HTTPException(
            status_code=403,
            detail="Not authorized to access this project's evaluations",
        )

    # Verify ground truth exists for this project
    groundtruth: models.GroundTruth | None = db.execute(
        select(models.GroundTruth).where(
            models.GroundTruth.project_id == project_id,
            models.GroundTruth.id == groundtruth_id,
        )
    ).scalar_one_or_none()

    if not groundtruth:
        raise HTTPException(status_code=404, detail="Ground truth not found")

    # Get all evaluations for this ground truth
    evaluations = list(
        db.execute(
            select(models.Evaluation).where(
                models.Evaluation.groundtruth_id == groundtruth_id
            )
        )
        .scalars()
        .all()
    )

    return [schemas.Evaluation.model_validate(eval) for eval in evaluations]


@router.get(
    "/{project_id}/evaluation/{evaluation_id}", response_model=schemas.EvaluationDetail
)
def get_evaluation_detail(
    *,
    db: Session = Depends(get_db),
    project_id: int,
    evaluation_id: int,
    current_user: models.User = Depends(get_current_user),
) -> schemas.EvaluationDetail:
    """Get detailed evaluation for a specific trial/ground truth pair."""
    project: models.Project | None = db.execute(
        select(models.Project).where(models.Project.id == project_id)
    ).scalar_one_or_none()

    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    if current_user.role != "admin" and project.owner_id != current_user.id:
        raise HTTPException(
            status_code=403,
            detail="Not authorized to access this project's evaluations",
        )

    evaluation: models.Evaluation | None = db.execute(
        select(models.Evaluation).where(models.Evaluation.id == evaluation_id)
    ).scalar_one_or_none()

    if not evaluation:
        raise HTTPException(status_code=404, detail="Evaluation not found")

    # Ensure the evaluation belongs to a trial in this project
    trial: models.Trial | None = db.execute(
        select(models.Trial).where(
            models.Trial.id == evaluation.trial_id,
            models.Trial.project_id == project_id,
        )
    ).scalar_one_or_none()

    if not trial:
        raise HTTPException(
            status_code=404, detail="Evaluation not found for this project"
        )

    # Get detailed evaluation data
    evaluation_detail = schemas.EvaluationDetail(
        id=evaluation.id,
        trial_id=evaluation.trial_id,
        groundtruth_id=evaluation.groundtruth_id,
        model=trial.llm_model,
        metrics=evaluation.metrics,
        document_count=len(evaluation.document_metrics),
        fields=evaluation.field_metrics,
        documents=evaluation.document_metrics,
        created_at=evaluation.created_at,
    )

    return evaluation_detail


@router.get("/{project_id}/groundtruth/{groundtruth_id}/preview", response_model=dict)
def preview_groundtruth(
    *,
    db: Session = Depends(get_db),
    project_id: int,
    groundtruth_id: int,
    limit: int = Query(10, description="Number of documents to preview"),
    current_user: models.User = Depends(get_current_user),
) -> dict:
    """Preview ground truth data and field structure."""
    project: models.Project | None = db.execute(
        select(models.Project).where(models.Project.id == project_id)
    ).scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    if current_user.role != "admin" and project.owner_id != current_user.id:
        raise HTTPException(
            status_code=403,
            detail="Not authorized to access this project's ground truth",
        )
    groundtruth: models.GroundTruth | None = db.execute(
        select(models.GroundTruth).where(
            models.GroundTruth.project_id == project_id,
            models.GroundTruth.id == groundtruth_id,
        )
    ).scalar_one_or_none()
    if not groundtruth:
        raise HTTPException(status_code=404, detail="Ground truth not found")
    # Load ground truth data
    from ....utils.evaluation import EvaluationEngine

    engine = EvaluationEngine(db)
    gt_data = engine._load_ground_truth(groundtruth)
    # Get field structure

    # For CSV/XLSX, also get available columns for ID selection
    available_columns = []
    if groundtruth.format in ["csv", "xlsx"]:
        # Load raw file to get all columns
        from ....dependencies import get_file

        content = get_file(groundtruth.file_uuid)

        if groundtruth.format == "csv":
            df = pd.read_csv(io.BytesIO(content))
        else:  # xlsx
            df = pd.read_excel(io.BytesIO(content))

        available_columns = df.columns.tolist()

    all_fields = set()
    field_types = {}
    sample_values = {}
    for doc_id, fields in list(gt_data.items())[:limit]:
        for field, value in fields.items():
            all_fields.add(field)
            # Infer field type
            if field not in field_types:
                if isinstance(value, bool):
                    field_types[field] = "boolean"
                elif isinstance(value, (int, float)):
                    field_types[field] = "number"
                elif isinstance(value, str):
                    # Check if it's a date
                    if re.match(r"\d{4}-\d{2}-\d{2}", value):
                        field_types[field] = "date"
                    else:
                        field_types[field] = "string"
                else:
                    field_types[field] = "string"
            # Collect sample values
            if field not in sample_values:
                sample_values[field] = []
            if value not in sample_values[field] and len(sample_values[field]) < 5:
                sample_values[field].append(value)
    return {
        "document_count": len(gt_data),
        "fields": list(all_fields),
        "field_types": field_types,
        "sample_values": sample_values,
        "preview_data": dict(list(gt_data.items())[:limit]),
        "available_columns": available_columns,
        "current_id_column": groundtruth.id_column_name,
    }


@router.post(
    "/{project_id}/groundtruth/{groundtruth_id}/mapping",
    response_model=schemas.GroundTruth,
)
def configure_field_mapping_legacy(
    *,
    db: Session = Depends(get_db),
    project_id: int,
    groundtruth_id: int,
    mappings: list[schemas.FieldMappingCreate] = Body(...),
    current_user: models.User = Depends(get_current_user),
) -> schemas.GroundTruth:
    """Configure field mappings for ground truth evaluation."""
    project: models.Project | None = db.execute(
        select(models.Project).where(models.Project.id == project_id)
    ).scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    if current_user.role != "admin" and project.owner_id != current_user.id:
        raise HTTPException(
            status_code=403, detail="Not authorized to configure ground truth mappings"
        )
    groundtruth: models.GroundTruth | None = db.execute(
        select(models.GroundTruth).where(
            models.GroundTruth.project_id == project_id,
            models.GroundTruth.id == groundtruth_id,
        )
    ).scalar_one_or_none()
    if not groundtruth:
        raise HTTPException(status_code=404, detail="Ground truth not found")
    # Delete existing mappings
    db.execute(
        delete(models.FieldMapping).where(
            models.FieldMapping.ground_truth_id == groundtruth_id
        )
    )
    # Create new mappings
    for mapping_data in mappings:
        mapping = models.FieldMapping(
            ground_truth_id=groundtruth_id, **mapping_data.model_dump()
        )
        db.add(mapping)
    # Invalidate existing evaluations
    db.execute(
        delete(models.Evaluation).where(
            models.Evaluation.groundtruth_id == groundtruth_id
        )
    )
    db.commit()
    db.refresh(groundtruth)
    return schemas.GroundTruth.model_validate(groundtruth)


@router.get(
    "/{project_id}/groundtruth/{groundtruth_id}/schema/{schema_id}/mapping",
    response_model=list[schemas.FieldMapping],
)
def get_field_mappings(
    *,
    db: Session = Depends(get_db),
    project_id: int,
    groundtruth_id: int,
    schema_id: int,
    current_user: models.User = Depends(get_current_user),
) -> list[schemas.FieldMapping]:
    """Get field mappings for a specific groundtruth-schema combination."""
    # Validate project access
    project: models.Project | None = db.execute(
        select(models.Project).where(models.Project.id == project_id)
    ).scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    if current_user.role != "admin" and project.owner_id != current_user.id:
        raise HTTPException(
            status_code=403, detail="Not authorized to access field mappings"
        )

    # Validate groundtruth exists
    groundtruth: models.GroundTruth | None = db.execute(
        select(models.GroundTruth).where(
            models.GroundTruth.project_id == project_id,
            models.GroundTruth.id == groundtruth_id,
        )
    ).scalar_one_or_none()
    if not groundtruth:
        raise HTTPException(status_code=404, detail="Ground truth not found")

    # Validate schema exists
    schema: models.Schema | None = db.execute(
        select(models.Schema).where(
            models.Schema.project_id == project_id,
            models.Schema.id == schema_id,
        )
    ).scalar_one_or_none()
    if not schema:
        raise HTTPException(status_code=404, detail="Schema not found")

    # Get mappings for this groundtruth-schema combination
    mappings = (
        db.execute(
            select(models.FieldMapping).where(
                models.FieldMapping.ground_truth_id == groundtruth_id,
                models.FieldMapping.schema_id == schema_id,
            )
        )
        .scalars()
        .all()
    )

    return [schemas.FieldMapping.model_validate(mapping) for mapping in mappings]


@router.delete(
    "/{project_id}/groundtruth/{groundtruth_id}/schema/{schema_id}/mapping",
    response_model=dict,
)
def delete_field_mappings(
    *,
    db: Session = Depends(get_db),
    project_id: int,
    groundtruth_id: int,
    schema_id: int,
    current_user: models.User = Depends(get_current_user),
) -> dict:
    """Delete all field mappings for a specific groundtruth-schema combination."""
    # Validate project access
    project: models.Project | None = db.execute(
        select(models.Project).where(models.Project.id == project_id)
    ).scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    if current_user.role != "admin" and project.owner_id != current_user.id:
        raise HTTPException(
            status_code=403, detail="Not authorized to delete field mappings"
        )

    # Validate groundtruth exists
    groundtruth: models.GroundTruth | None = db.execute(
        select(models.GroundTruth).where(
            models.GroundTruth.project_id == project_id,
            models.GroundTruth.id == groundtruth_id,
        )
    ).scalar_one_or_none()
    if not groundtruth:
        raise HTTPException(status_code=404, detail="Ground truth not found")

    # Validate schema exists
    schema: models.Schema | None = db.execute(
        select(models.Schema).where(
            models.Schema.project_id == project_id,
            models.Schema.id == schema_id,
        )
    ).scalar_one_or_none()
    if not schema:
        raise HTTPException(status_code=404, detail="Schema not found")

    # Delete mappings for this groundtruth-schema combination
    deleted_count = db.execute(
        delete(models.FieldMapping).where(
            models.FieldMapping.ground_truth_id == groundtruth_id,
            models.FieldMapping.schema_id == schema_id,
        )
    ).rowcount

    # Invalidate related evaluations
    db.execute(
        delete(models.Evaluation).where(
            models.Evaluation.groundtruth_id == groundtruth_id,
            models.Evaluation.trial_id.in_(
                select(models.Trial.id).where(models.Trial.schema_id == schema_id)
            ),
        )
    )

    db.commit()

    return {"deleted_mappings": deleted_count}


@router.get(
    "/{project_id}/groundtruth/{groundtruth_id}/schema/{schema_id}/mapping/suggest",
    response_model=list[schemas.FieldMapping],
)
def suggest_field_mappings(
    *,
    db: Session = Depends(get_db),
    project_id: int,
    groundtruth_id: int,
    schema_id: int,
    current_user: models.User = Depends(get_current_user),
) -> list[schemas.FieldMapping]:
    """Suggest field mappings based on schema and ground truth."""
    # Validate project access
    project: models.Project | None = db.execute(
        select(models.Project).where(models.Project.id == project_id)
    ).scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    if current_user.role != "admin" and project.owner_id != current_user.id:
        raise HTTPException(
            status_code=403, detail="Not authorized to access this project"
        )

    # Get schema
    schema: models.Schema | None = db.execute(
        select(models.Schema).where(
            models.Schema.project_id == project_id, models.Schema.id == schema_id
        )
    ).scalar_one_or_none()
    if not schema:
        raise HTTPException(status_code=404, detail="Schema not found")

    # Get ground truth
    groundtruth: models.GroundTruth | None = db.execute(
        select(models.GroundTruth).where(
            models.GroundTruth.project_id == project_id,
            models.GroundTruth.id == groundtruth_id,
        )
    ).scalar_one_or_none()
    if not groundtruth:
        raise HTTPException(status_code=404, detail="Ground truth not found")

    # Load ground truth data
    from ....utils.evaluation import EvaluationEngine

    engine = EvaluationEngine(db)
    gt_data = engine._load_ground_truth(groundtruth)

    # Extract schema fields
    schema_fields = {}
    extract_field_types_from_schema(schema.schema_definition, schema_fields)

    # Get ground truth fields
    gt_fields = set()
    for fields in gt_data.values():
        gt_fields.update(fields.keys())

    # Suggest mappings
    suggestions = []
    for schema_field, field_type in schema_fields.items():
        # Try exact match
        if schema_field in gt_fields:
            suggestions.append(
                {
                    "id": 0,
                    "ground_truth_id": groundtruth_id,
                    "schema_id": schema_id,
                    "schema_field": schema_field,
                    "ground_truth_field": schema_field,
                    "field_type": field_type,
                    "comparison_method": _get_comparison_method(field_type),
                    "confidence": 1.0,
                }
            )
            continue

        # Try fuzzy matching
        best_match = None
        best_score = 0
        for gt_field in gt_fields:
            score = fuzz.ratio(schema_field.lower(), gt_field.lower())
            if score > best_score and score > 70:
                best_score = score
                best_match = gt_field

        if best_match:
            suggestions.append(
                {
                    "id": 0,
                    "ground_truth_id": groundtruth_id,
                    "schema_id": schema_id,
                    "schema_field": schema_field,
                    "ground_truth_field": best_match,
                    "field_type": field_type,
                    "comparison_method": _get_comparison_method(field_type),
                    "confidence": best_score / 100.0,
                }
            )

    return suggestions


def _get_comparison_method(field_type: str) -> str:
    """Get default comparison method for field type."""
    type_to_method = {
        "boolean": "boolean",
        "number": "numeric",
        "date": "date",
        "category": "category",
        "string": "exact",
    }
    return type_to_method.get(field_type, "exact")


@router.post(
    "/{project_id}/groundtruth/{groundtruth_id}/schema/{schema_id}/validate-json",
    response_model=dict,
)
def validate_json_ground_truth(
    *,
    db: Session = Depends(get_db),
    project_id: int,
    groundtruth_id: int,
    schema_id: int,
    current_user: models.User = Depends(get_current_user),
) -> dict:
    """Validate JSON ground truth against schema definition."""
    # ... access checks ...

    groundtruth = db.get(models.GroundTruth, groundtruth_id)
    schema = db.get(models.Schema, schema_id)

    # Only for JSON format
    if groundtruth.format not in ["json", "zip"]:
        return {"errors": [], "warnings": ["Validation only applies to JSON format"]}

    # Load ground truth data
    from ....utils.evaluation import EvaluationEngine

    engine = EvaluationEngine(db)
    gt_data = engine._load_ground_truth(groundtruth)

    errors = []
    warnings = []

    from ....utils.helpers import extract_required_fields_from_schema, check_missing_fields_nested, check_field_types, find_extra_fields

    # Validate structure matches schema
    schema_def = schema.schema_definition
    required_fields = extract_required_fields_from_schema(schema_def)

    # Check a sample of documents
    sample_size = min(10, len(gt_data))
    for i, (doc_id, doc_data) in enumerate(list(gt_data.items())[:sample_size]):
        # Check required fields
        missing = check_missing_fields_nested(doc_data, required_fields, "")
        if missing:
            errors.extend([f"Document {doc_id}: Missing {field}" for field in missing])

        # Check data types
        type_errors = check_field_types(doc_data, schema_def, "")
        if type_errors:
            errors.extend([f"Document {doc_id}: {err}" for err in type_errors])

    # Extra fields are warnings
    extra = find_extra_fields(doc_data, schema_def)
    if extra:
        warnings.extend([f"Extra field '{field}' not in schema" for field in extra])

    return {"errors": errors[:10], "warnings": warnings[:10]}  # Limit to 10 each


@router.post(
    "/{project_id}/trial/{trial_id}/evaluate", response_model=schemas.EvaluationSummary
)
def evaluate_trial(
    *,
    db: Session = Depends(get_db),
    project_id: int,
    trial_id: int,
    groundtruth_id: int = Query(...),
    force_recalculate: bool = Query(False),
    current_user: models.User = Depends(get_current_user),
) -> schemas.EvaluationSummary:
    """Evaluate a trial against ground truth with comprehensive error handling."""

    # Validate project access
    project: models.Project | None = db.execute(
        select(models.Project).where(models.Project.id == project_id)
    ).scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    if current_user.role != "admin" and project.owner_id != current_user.id:
        raise HTTPException(
            status_code=403, detail="Not authorized to evaluate trials for this project"
        )

    # Verify trial exists and belongs to project
    trial: models.Trial | None = db.execute(
        select(models.Trial).where(
            models.Trial.project_id == project_id, models.Trial.id == trial_id
        )
    ).scalar_one_or_none()
    if not trial:
        raise HTTPException(status_code=404, detail="Trial not found")

    # Verify ground truth exists and belongs to project
    groundtruth: models.GroundTruth | None = db.execute(
        select(models.GroundTruth).where(
            models.GroundTruth.project_id == project_id,
            models.GroundTruth.id == groundtruth_id,
        )
    ).scalar_one_or_none()
    if not groundtruth:
        raise HTTPException(status_code=404, detail="Ground truth not found")

    # Pre-validation checks
    from ....utils.evaluation import EvaluationEngine

    engine = EvaluationEngine(db)

    try:
        validation_result = engine._validate_evaluation_prerequisites(
            trial_id, groundtruth_id
        )
        if not validation_result["valid"]:
            # Provide detailed error information
            error_details = {
                "message": "Cannot evaluate trial due to validation errors",
                "errors": validation_result["errors"],
                "suggestions": [],
            }

            # Add specific suggestions based on error types
            for error in validation_result["errors"]:
                if "No field mappings configured" in error:
                    error_details["suggestions"].append(
                        "Configure field mappings between your ground truth data and schema fields"
                    )
                elif "No results found" in error:
                    error_details["suggestions"].append(
                        "Ensure the trial has completed successfully and produced results"
                    )
                elif "documents have matching ground truth" in error:
                    error_details["suggestions"].append(
                        "Check that your ground truth file contains keys that match your document IDs or filenames"
                    )

            raise HTTPException(status_code=400, detail=error_details)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Validation failed: {str(e)}")

    # Run evaluation
    try:
        evaluation = engine.evaluate_trial(
            trial_id=trial_id,
            groundtruth_id=groundtruth_id,
            force_recalculate=force_recalculate,
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Evaluation failed: {str(e)}")

    # Build enhanced summary with error information
    field_summaries = []
    for field_name, metrics in evaluation.field_metrics.items():
        # Calculate error count
        error_count = sum(metrics.get("error_distribution", {}).values())

        # Get sample errors
        sample_errors = []
        error_details = (
            db.query(models.EvaluationMetric)
            .filter(
                models.EvaluationMetric.evaluation_id == evaluation.id,
                models.EvaluationMetric.field_name == field_name,
                ~models.EvaluationMetric.is_correct,
            )
            .limit(5)
            .all()
        )
        for detail in error_details:
            sample_errors.append(
                {
                    "document_id": detail.document_id,
                    "ground_truth": detail.ground_truth_value,
                    "predicted": detail.predicted_value,
                    "error_type": detail.error_type,
                }
            )

        field_summaries.append(
            schemas.FieldEvaluationSummary(
                field_name=field_name,
                accuracy=metrics.get("accuracy", 0),
                total_count=metrics.get("total_count", 0),
                correct_count=metrics.get("correct_count", 0),
                error_distribution=metrics.get("error_distribution", {}),
                sample_errors=sample_errors,
                error_count=error_count,
            )
        )

    # Build document summaries with enhanced error information
    document_summaries = []
    total_errors = 0
    error_documents = []

    for doc_metrics in evaluation.document_metrics:
        doc_id = doc_metrics["document_id"]

        # Check if this document had errors
        has_error = "error" in doc_metrics
        if has_error:
            total_errors += 1
            error_documents.append(doc_id)

        # Get document name
        document = db.get(models.Document, doc_id)
        document_name = None
        if document and document.original_file:
            document_name = document.original_file.file_name

        # Get field details
        field_details = {}
        details = (
            db.query(models.EvaluationMetric)
            .filter(
                models.EvaluationMetric.evaluation_id == evaluation.id,
                models.EvaluationMetric.document_id == doc_id,
            )
            .all()
        )
        for detail in details:
            field_details[detail.field_name] = schemas.EvaluationMetricDetail(
                document_id=doc_id,
                field_name=detail.field_name,
                ground_truth_value=detail.ground_truth_value,
                predicted_value=detail.predicted_value,
                is_correct=detail.is_correct,
                error_type=detail.error_type,
                confidence_score=detail.confidence_score,
            )

        document_summary = schemas.DocumentEvaluationDetail(
            document_id=doc_id,
            accuracy=doc_metrics.get("accuracy", 0.0),
            correct_fields=doc_metrics.get("correct_fields", 0),
            total_fields=doc_metrics.get("total_fields", 0),
            missing_fields=doc_metrics.get("missing_fields", []),
            incorrect_fields=doc_metrics.get("incorrect_fields", []),
            field_details=field_details,
            has_error=has_error,
            document_name=document_name,
        )

        # Add error information if present
        if has_error:
            document_summary.error = doc_metrics["error"]

        document_summaries.append(document_summary)

    return schemas.EvaluationSummary(
        id=evaluation.id,
        trial_id=evaluation.trial_id,
        groundtruth_id=evaluation.groundtruth_id,
        overall_metrics=evaluation.metrics,
        field_summaries=field_summaries,
        document_summaries=document_summaries,
        confusion_matrices=evaluation.confusion_matrices,
        created_at=evaluation.created_at,
        total_errors=total_errors,
        error_documents=error_documents,
    )


@router.get(
    "/{project_id}/evaluation/{evaluation_id}/document/{document_id}",
    response_model=schemas.DocumentEvaluationDetail,
)
def get_document_evaluation(
    *,
    db: Session = Depends(get_db),
    project_id: int,
    evaluation_id: int,
    document_id: int,
    current_user: models.User = Depends(get_current_user),
) -> schemas.DocumentEvaluationDetail:
    """Get detailed evaluation for a single document."""
    project: models.Project | None = db.execute(
        select(models.Project).where(models.Project.id == project_id)
    ).scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    if current_user.role != "admin" and project.owner_id != current_user.id:
        raise HTTPException(
            status_code=403,
            detail="Not authorized to access this project's evaluations",
        )
    # Get evaluation
    evaluation: models.Evaluation | None = db.execute(
        select(models.Evaluation).where(models.Evaluation.id == evaluation_id)
    ).scalar_one_or_none()
    if not evaluation:
        raise HTTPException(status_code=404, detail="Evaluation not found")
    # Verify evaluation belongs to project
    trial: models.Trial | None = db.execute(
        select(models.Trial).where(
            models.Trial.id == evaluation.trial_id,
            models.Trial.project_id == project_id,
        )
    ).scalar_one_or_none()
    if not trial:
        raise HTTPException(
            status_code=404, detail="Evaluation not found for this project"
        )
    # Find document metrics
    doc_metrics = None
    for metrics in evaluation.document_metrics:
        if metrics["document_id"] == document_id:
            doc_metrics = metrics
            break
    if not doc_metrics:
        raise HTTPException(status_code=404, detail="Document not found in evaluation")
    # Get detailed metrics
    field_details = {}
    details = (
        db.query(models.EvaluationMetric)
        .filter(
            models.EvaluationMetric.evaluation_id == evaluation_id,
            models.EvaluationMetric.document_id == document_id,
        )
        .all()
    )
    for detail in details:
        field_details[detail.field_name] = schemas.EvaluationMetricDetail(
            document_id=document_id,
            field_name=detail.field_name,
            ground_truth_value=detail.ground_truth_value,
            predicted_value=detail.predicted_value,
            is_correct=detail.is_correct,
            error_type=detail.error_type,
            confidence_score=detail.confidence_score,
        )
    return schemas.DocumentEvaluationDetail(
        document_id=document_id,
        accuracy=doc_metrics["accuracy"],
        correct_fields=doc_metrics["correct_fields"],
        total_fields=doc_metrics["total_fields"],
        missing_fields=doc_metrics.get("missing_fields", []),
        incorrect_fields=doc_metrics.get("incorrect_fields", []),
        field_details=field_details,
    )


@router.get("/{project_id}/evaluations/download", response_class=Response)
def download_evaluations_report(
    *,
    db: Session = Depends(get_db),
    project_id: int,
    evaluation_ids: str = Query(...),
    format: str = Query("csv", enum=["csv", "xlsx"]),
    include_details: bool = Query(True),
    current_user: models.User = Depends(get_current_user),
) -> Response:
    """Download evaluation report in CSV or Excel format."""
    try:
        evaluation_ids_list = [int(i) for i in evaluation_ids.split(",")]
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid evaluation_ids")
    project: models.Project | None = db.execute(
        select(models.Project).where(models.Project.id == project_id)
    ).scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    if current_user.role != "admin" and project.owner_id != current_user.id:
        raise HTTPException(
            status_code=403,
            detail="Not authorized to access this project's evaluations",
        )
    # Get evaluations
    evaluations = []
    for eval_id in evaluation_ids_list:
        eval_obj = db.execute(
            select(models.Evaluation).where(models.Evaluation.id == eval_id)
        ).scalar_one_or_none()
        if eval_obj:
            # Verify it belongs to project
            trial = db.execute(
                select(models.Trial).where(
                    models.Trial.id == eval_obj.trial_id,
                    models.Trial.project_id == project_id,
                )
            ).scalar_one_or_none()
            if trial:
                evaluations.append((eval_obj, trial))
    if not evaluations:
        raise HTTPException(status_code=404, detail="No evaluations found")
    # Create report data
    if format == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        # Summary sheet
        writer.writerow(
            [
                "Evaluation ID",
                "Trial ID",
                "Model",
                "Ground Truth",
                "Accuracy",
                "Precision",
                "Recall",
                "F1 Score",
                "Total Documents",
                "Total Fields",
                "Created At",
            ]
        )
        for eval, trial in evaluations:
            gt = db.query(models.GroundTruth).get(eval.groundtruth_id)
            writer.writerow(
                [
                    eval.id,
                    eval.trial_id,
                    trial.llm_model,
                    gt.name if gt else "Unknown",
                    eval.metrics.get("accuracy", 0),
                    eval.metrics.get("precision", 0),
                    eval.metrics.get("recall", 0),
                    eval.metrics.get("f1_score", 0),
                    eval.metrics.get("total_documents", 0),
                    eval.metrics.get("total_fields", 0),
                    eval.created_at.isoformat(),
                ]
            )
        if include_details:
            writer.writerow([])  # Empty row
            writer.writerow(["Field-Level Metrics"])
            writer.writerow(
                [
                    "Evaluation ID",
                    "Field Name",
                    "Accuracy",
                    "Total Count",
                    "Correct Count",
                ]
            )
            for eval, _ in evaluations:
                for field, metrics in eval.field_metrics.items():
                    writer.writerow(
                        [
                            eval.id,
                            field,
                            metrics.get("accuracy", 0),
                            metrics.get("total_count", 0),
                            metrics.get("correct_count", 0),
                        ]
                    )
        content = output.getvalue()
        media_type = "text/csv"
        filename = f"evaluation_report_{project_id}.csv"
    else:  # xlsx
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
            # Summary sheet
            summary_data = []
            for eval, trial in evaluations:
                gt = db.query(models.GroundTruth).get(eval.groundtruth_id)
                summary_data.append(
                    {
                        "Evaluation ID": eval.id,
                        "Trial ID": eval.trial_id,
                        "Model": trial.llm_model,
                        "Ground Truth": gt.name if gt else "Unknown",
                        "Accuracy": eval.metrics.get("accuracy", 0),
                        "Precision": eval.metrics.get("precision", 0),
                        "Recall": eval.metrics.get("recall", 0),
                        "F1 Score": eval.metrics.get("f1_score", 0),
                        "Total Documents": eval.metrics.get("total_documents", 0),
                        "Total Fields": eval.metrics.get("total_fields", 0),
                        "Created At": eval.created_at.isoformat(),
                    }
                )
            pd.DataFrame(summary_data).to_excel(
                writer, sheet_name="Summary", index=False
            )
            if include_details:
                # Field metrics sheet
                field_data = []
                for eval, _ in evaluations:
                    for field, metrics in eval.field_metrics.items():
                        field_data.append(
                            {
                                "Evaluation ID": eval.id,
                                "Field Name": field,
                                "Accuracy": metrics.get("accuracy", 0),
                                "Total Count": metrics.get("total_count", 0),
                                "Correct Count": metrics.get("correct_count", 0),
                            }
                        )
                pd.DataFrame(field_data).to_excel(
                    writer, sheet_name="Field Metrics", index=False
                )
                # Document metrics sheet
                doc_data = []
                for eval, _ in evaluations:
                    for doc_metrics in eval.document_metrics:
                        doc_data.append(
                            {
                                "Evaluation ID": eval.id,
                                "Document ID": doc_metrics["document_id"],
                                "Accuracy": doc_metrics["accuracy"],
                                "Correct Fields": doc_metrics["correct_fields"],
                                "Total Fields": doc_metrics["total_fields"],
                            }
                        )
                pd.DataFrame(doc_data).to_excel(
                    writer, sheet_name="Document Metrics", index=False
                )
        output.seek(0)
        content = output.getvalue()
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        filename = f"evaluation_report_{project_id}.xlsx"
    return Response(
        content=content,
        media_type=media_type,
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.post(
    "/{project_id}/evaluation/batch", response_model=list[schemas.EvaluationSummary]
)
def batch_evaluate_trials(
    *,
    db: Session = Depends(get_db),
    project_id: int,
    trial_ids: list[int] = Body(...),
    groundtruth_id: int = Body(...),
    force_recalculate: bool = Body(False),
    current_user: models.User = Depends(get_current_user),
) -> list[schemas.EvaluationSummary]:
    """Evaluate multiple trials against ground truth."""
    project: models.Project | None = db.execute(
        select(models.Project).where(models.Project.id == project_id)
    ).scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    if current_user.role != "admin" and project.owner_id != current_user.id:
        raise HTTPException(
            status_code=403, detail="Not authorized to evaluate trials for this project"
        )
    # Verify ground truth
    groundtruth: models.GroundTruth | None = db.execute(
        select(models.GroundTruth).where(
            models.GroundTruth.project_id == project_id,
            models.GroundTruth.id == groundtruth_id,
        )
    ).scalar_one_or_none()
    if not groundtruth:
        raise HTTPException(status_code=404, detail="Ground truth not found")
    # Evaluate trials
    from ....utils.evaluation import EvaluationEngine

    engine = EvaluationEngine(db)
    results = []
    errors = []
    for trial_id in trial_ids:
        try:
            # Verify trial exists
            trial: models.Trial | None = db.execute(
                select(models.Trial).where(
                    models.Trial.project_id == project_id, models.Trial.id == trial_id
                )
            ).scalar_one_or_none()
            if not trial:
                errors.append(f"Trial {trial_id} not found")
                continue
            evaluation = engine.evaluate_trial(
                trial_id=trial_id,
                groundtruth_id=groundtruth_id,
                force_recalculate=force_recalculate,
            )
            # Build summary (same as single evaluation)
            # ... (same summary building code as above)
            results.append(evaluation)
        except Exception as e:
            errors.append(f"Error evaluating trial {trial_id}: {str(e)}")
    if errors and not results:
        raise HTTPException(
            status_code=400, detail=f"All evaluations failed: {'; '.join(errors)}"
        )
    return results


@router.get("/{project_id}/evaluation/compare", response_model=dict)
def compare_evaluations(
    *,
    db: Session = Depends(get_db),
    project_id: int,
    evaluation_ids: list[int] = Query(...),
    current_user: models.User = Depends(get_current_user),
) -> dict:
    """Compare multiple evaluations side by side."""

    project: models.Project | None = db.execute(
        select(models.Project).where(models.Project.id == project_id)
    ).scalar_one_or_none()

    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    if current_user.role != "admin" and project.owner_id != current_user.id:
        raise HTTPException(
            status_code=403,
            detail="Not authorized to access this project's evaluations",
        )

    # Get evaluations
    evaluations = []
    for eval_id in evaluation_ids:
        eval_obj = db.execute(
            select(models.Evaluation).where(models.Evaluation.id == eval_id)
        ).scalar_one_or_none()

        if eval_obj:
            # Verify it belongs to project
            trial = db.execute(
                select(models.Trial).where(
                    models.Trial.id == eval_obj.trial_id,
                    models.Trial.project_id == project_id,
                )
            ).scalar_one_or_none()

            if trial:
                evaluations.append({"evaluation": eval_obj, "trial": trial})

    if not evaluations:
        raise HTTPException(status_code=404, detail="No evaluations found")

    # Build comparison
    comparison = {
        "evaluations": [],
        "overall_comparison": {},
        "field_comparison": {},
        "model_comparison": {},
    }

    # Collect all fields
    all_fields = set()
    for eval_data in evaluations:
        eval = eval_data["evaluation"]
        all_fields.update(eval.field_metrics.keys())

    # Build evaluation summaries
    for eval_data in evaluations:
        eval = eval_data["evaluation"]
        trial = eval_data["trial"]

        comparison["evaluations"].append(
            {
                "id": eval.id,
                "trial_id": eval.trial_id,
                "model": trial.llm_model,
                "groundtruth_id": eval.groundtruth_id,
                "metrics": eval.metrics,
                "created_at": eval.created_at.isoformat(),
            }
        )

    # Compare overall metrics
    metrics_to_compare = ["accuracy", "precision", "recall", "f1_score"]
    for metric in metrics_to_compare:
        comparison["overall_comparison"][metric] = []
        for eval_data in evaluations:
            eval = eval_data["evaluation"]
            trial = eval_data["trial"]
            comparison["overall_comparison"][metric].append(
                {
                    "evaluation_id": eval.id,
                    "model": trial.llm_model,
                    "value": eval.metrics.get(metric, 0),
                }
            )

    # Compare field metrics
    for field in sorted(all_fields):
        comparison["field_comparison"][field] = []
        for eval_data in evaluations:
            eval = eval_data["evaluation"]
            trial = eval_data["trial"]
            field_metric = eval.field_metrics.get(field, {})
            comparison["field_comparison"][field].append(
                {
                    "evaluation_id": eval.id,
                    "model": trial.llm_model,
                    "accuracy": field_metric.get("accuracy", 0),
                    "total_count": field_metric.get("total_count", 0),
                    "correct_count": field_metric.get("correct_count", 0),
                }
            )

    # Group by model
    model_groups = {}
    for eval_data in evaluations:
        eval = eval_data["evaluation"]
        trial = eval_data["trial"]
        model = trial.llm_model

        if model not in model_groups:
            model_groups[model] = []
        model_groups[model].append(eval.metrics.get("accuracy", 0))

    comparison["model_comparison"] = {
        model: {
            "average_accuracy": sum(accuracies) / len(accuracies),
            "evaluation_count": len(accuracies),
            "min_accuracy": min(accuracies),
            "max_accuracy": max(accuracies),
        }
        for model, accuracies in model_groups.items()
    }

    return comparison


@router.get(
    "/{project_id}/evaluation/{evaluation_id}/errors", response_model=list[dict]
)
def get_evaluation_errors(
    *,
    db: Session = Depends(get_db),
    project_id: int,
    evaluation_id: int,
    field_name: str | None = Query(None),
    error_type: str | None = Query(None),
    limit: int = Query(100),
    current_user: models.User = Depends(get_current_user),
) -> list[dict]:
    """Get detailed error analysis for an evaluation."""

    project: models.Project | None = db.execute(
        select(models.Project).where(models.Project.id == project_id)
    ).scalar_one_or_none()

    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    if current_user.role != "admin" and project.owner_id != current_user.id:
        raise HTTPException(
            status_code=403,
            detail="Not authorized to access this project's evaluations",
        )

    # Get evaluation
    evaluation: models.Evaluation | None = db.execute(
        select(models.Evaluation).where(models.Evaluation.id == evaluation_id)
    ).scalar_one_or_none()

    if not evaluation:
        raise HTTPException(status_code=404, detail="Evaluation not found")

    # Verify evaluation belongs to project
    trial: models.Trial | None = db.execute(
        select(models.Trial).where(
            models.Trial.id == evaluation.trial_id,
            models.Trial.project_id == project_id,
        )
    ).scalar_one_or_none()

    if not trial:
        raise HTTPException(
            status_code=404, detail="Evaluation not found for this project"
        )

    # Get errors from detailed metrics
    query = db.query(models.EvaluationMetric).filter(
        models.EvaluationMetric.evaluation_id == evaluation_id,
        ~models.EvaluationMetric.is_correct,
    )

    if field_name:
        query = query.filter(models.EvaluationMetric.field_name == field_name)

    if error_type:
        query = query.filter(models.EvaluationMetric.error_type == error_type)

    errors = query.limit(limit).all()

    # Build error details
    error_details = []
    for error in errors:
        # Get document info
        document = db.query(models.Document).get(error.document_id)

        error_detail = {
            "document_id": error.document_id,
            "document_name": document.original_file.file_name
            if document and document.original_file
            else "Unknown",
            "field_name": error.field_name,
            "ground_truth_value": error.ground_truth_value,
            "predicted_value": error.predicted_value,
            "error_type": error.error_type,
            "confidence_score": error.confidence_score,
        }

        # Add context if available
        if document:
            # Extract surrounding text for context
            text_snippet = document.text[:500] if document.text else ""
            error_detail["context"] = text_snippet

        error_details.append(error_detail)

    return error_details


@router.post(
    "/{project_id}/groundtruth/{groundtruth_id}/schema/{schema_id}/auto-map",
    response_model=dict,
)
def auto_map_fields(
    *,
    db: Session = Depends(get_db),
    project_id: int,
    groundtruth_id: int,
    schema_id: int,
    confidence_threshold: float = Body(0.7),
    current_user: models.User = Depends(get_current_user),
) -> dict:
    """Automatically map ground truth fields to schema fields."""
    # Validate project access
    project: models.Project | None = db.execute(
        select(models.Project).where(models.Project.id == project_id)
    ).scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    if current_user.role != "admin" and project.owner_id != current_user.id:
        raise HTTPException(
            status_code=403, detail="Not authorized to access this project"
        )

    # Get schema and ground truth
    schema: models.Schema | None = db.execute(
        select(models.Schema).where(
            models.Schema.project_id == project_id, models.Schema.id == schema_id
        )
    ).scalar_one_or_none()
    if not schema:
        raise HTTPException(status_code=404, detail="Schema not found")

    groundtruth: models.GroundTruth | None = db.execute(
        select(models.GroundTruth).where(
            models.GroundTruth.project_id == project_id,
            models.GroundTruth.id == groundtruth_id,
        )
    ).scalar_one_or_none()
    if not groundtruth:
        raise HTTPException(status_code=404, detail="Ground truth not found")

    # Load ground truth data
    from ....utils.evaluation import EvaluationEngine

    engine = EvaluationEngine(db)
    gt_data = engine._load_ground_truth(groundtruth)

    # Extract schema fields
    schema_fields = {}
    extract_field_types_from_schema(schema.schema_definition, schema_fields)

    # Get ground truth fields and sample values
    gt_fields = {}
    for doc_id, fields in list(gt_data.items())[:10]:  # Sample first 10 docs
        for field, value in fields.items():
            if field not in gt_fields:
                gt_fields[field] = []
            if value not in gt_fields[field] and len(gt_fields[field]) < 5:
                gt_fields[field].append(value)

    # Auto-map fields
    from ....utils.field_mapping import FieldMapper

    mapper = FieldMapper()
    mappings = mapper.auto_map(
        schema_fields=schema_fields,
        ground_truth_fields=gt_fields,
        confidence_threshold=confidence_threshold,
    )

    # Delete existing mappings for this groundtruth-schema combination
    db.execute(
        delete(models.FieldMapping).where(
            models.FieldMapping.ground_truth_id == groundtruth_id,
            models.FieldMapping.schema_id == schema_id,
        )
    )

    # Apply mappings
    applied_count = 0
    for mapping in mappings:
        if mapping["confidence"] >= confidence_threshold:
            field_mapping = models.FieldMapping(
                ground_truth_id=groundtruth_id,
                schema_id=schema_id,
                schema_field=mapping["schema_field"],
                ground_truth_field=mapping["ground_truth_field"],
                field_type=mapping["field_type"],
                comparison_method=mapping["comparison_method"],
                comparison_options=mapping.get("comparison_options", {}),
            )
            db.add(field_mapping)
            applied_count += 1

    # Invalidate existing evaluations for this groundtruth-schema combination
    db.execute(
        delete(models.Evaluation).where(
            models.Evaluation.groundtruth_id == groundtruth_id,
            models.Evaluation.trial_id.in_(
                select(models.Trial.id).where(models.Trial.schema_id == schema_id)
            ),
        )
    )

    db.commit()

    return {
        "total_schema_fields": len(schema_fields),
        "total_ground_truth_fields": len(gt_fields),
        "suggested_mappings": len(mappings),
        "applied_mappings": applied_count,
        "mappings": mappings,
    }


@router.get(
    "/{project_id}/groundtruth/{groundtruth_id}/schema/{schema_id}/mapping/status",
    response_model=dict,
)
def check_mapping_status(
    *,
    db: Session = Depends(get_db),
    project_id: int,
    groundtruth_id: int,
    schema_id: int,
    current_user: models.User = Depends(get_current_user),
) -> dict:
    """Check if field mappings exist for a groundtruth-schema combination."""
    # Validate project access
    project: models.Project | None = db.execute(
        select(models.Project).where(models.Project.id == project_id)
    ).scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    if current_user.role != "admin" and project.owner_id != current_user.id:
        raise HTTPException(
            status_code=403, detail="Not authorized to access this project"
        )

    # Validate groundtruth and schema exist
    groundtruth: models.GroundTruth | None = db.execute(
        select(models.GroundTruth).where(
            models.GroundTruth.project_id == project_id,
            models.GroundTruth.id == groundtruth_id,
        )
    ).scalar_one_or_none()
    if not groundtruth:
        raise HTTPException(status_code=404, detail="Ground truth not found")

    schema: models.Schema | None = db.execute(
        select(models.Schema).where(
            models.Schema.project_id == project_id,
            models.Schema.id == schema_id,
        )
    ).scalar_one_or_none()
    if not schema:
        raise HTTPException(status_code=404, detail="Schema not found")

    # Check mapping status
    mapping_count = db.execute(
        select(func.count(models.FieldMapping.id)).where(
            models.FieldMapping.ground_truth_id == groundtruth_id,
            models.FieldMapping.schema_id == schema_id,
        )
    ).scalar()

    # Get schema field count
    schema_fields = {}
    extract_field_types_from_schema(schema.schema_definition, schema_fields)
    schema_field_count = len(schema_fields)

    return {
        "has_mappings": mapping_count > 0,
        "mapping_count": mapping_count,
        "schema_field_count": schema_field_count,
        "mapping_complete": mapping_count == schema_field_count,
        "groundtruth_name": groundtruth.name,
        "schema_name": schema.schema_name,
    }
