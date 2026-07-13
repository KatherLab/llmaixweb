# backend/src/routers/v1/endpoints/files.py
"""File management endpoints for projects."""

import datetime
import io
import json
import logging
import zipfile

from fastapi import (
    APIRouter,
    Body,
    Depends,
    File,
    Form,
    HTTPException,
    Query,
    UploadFile,
)
from fastapi.responses import Response, StreamingResponse
from pydantic import ValidationError
from sqlalchemy import and_, distinct, func, or_, select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session, selectinload

from .... import models, schemas
from ....core.security import (
    admin_has_global_project_access,
    can_access_project,
    get_current_user,
)
from ....dependencies import (
    get_db,
    get_file,
    hash_measure_and_head,
    remove_file,
    save_upload_stream,
    stream_file,
)
from ....middleware.error_handlers import (
    internal_error_message,
    record_internal_error,
)
from ....models.project import document_set_association
from ....utils.audit import record_audit
from ....utils.deletion import (
    cascade_clear_document_references,
    compute_document_dependencies,
)
from ....utils.enums import AuditAction, FileCreator
from ....utils.helpers import detect_structured_mime

logger = logging.getLogger(__name__)

router = APIRouter()


def check_project_access(
    project_id: int, current_user: models.User, db: Session, permission: str = "read"
) -> models.Project:
    """Check if user has access to project."""
    project = db.execute(
        select(models.Project).where(models.Project.id == project_id)
    ).scalar_one_or_none()

    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    # Admin has full access only when cross-user project access is enabled
    if admin_has_global_project_access(current_user):
        return project

    # Owner has full access
    if project.owner_id == current_user.id:
        return project

    # For non-owners, check specific permissions if needed
    # You could extend this with a project_members table for shared projects
    raise HTTPException(
        status_code=403, detail=f"Not authorized to {permission} this project"
    )


# Statuses matched exactly against the latest preprocessing task's status.
# ("processing" is a synthetic bucket that also includes pending/in_progress.)
_EXACT_FILE_STATUSES = frozenset({"pending", "in_progress", "completed", "failed"})
# "processing" expands to these underlying statuses.
_PROCESSING_STATUSES = ["pending", "processing", "in_progress"]
# Files with no preprocessing task at all.
_NO_TASK_STATUSES = frozenset({"not_preprocessed", "notstarted", "none"})


def _apply_file_status_filter(query, status: str, latest_task_status_with_status):
    """Apply the preprocessing-status filter to a file query.

    Shared by the page query and the count query so they can't diverge (they
    previously did: the count branch missed ``in_progress``, so filtering by
    ``status=in_progress`` returned the right rows but the wrong total).
    """
    status_lower = status.lower()
    if status_lower == "processing":
        return query.where(
            latest_task_status_with_status.c.latest_status.in_(_PROCESSING_STATUSES)
        )
    if status_lower in _EXACT_FILE_STATUSES:
        return query.where(
            latest_task_status_with_status.c.latest_status == status_lower
        )
    if status_lower in _NO_TASK_STATUSES:
        return query.where(latest_task_status_with_status.c.latest_status.is_(None))
    # Unknown status: no filter (matches prior behavior).
    return query


@router.get("", response_model=schemas.PaginatedFiles)
def get_project_files(
    *,
    db: Session = Depends(get_db),
    project_id: int,
    # Filter parameters
    search: str | None = Query(None, description="Search in filename"),
    file_type: str | None = Query(None, description="Filter by file type (MIME type)"),
    status: str | None = Query(
        None,
        description="Filter by preprocessing status (pending, processing, completed, failed)",
    ),
    file_creator: FileCreator | None = Query(None),
    date_from: datetime.datetime | None = Query(None),
    date_to: datetime.datetime | None = Query(None),
    min_size: int | None = Query(None, description="Minimum file size in bytes"),
    max_size: int | None = Query(None, description="Maximum file size in bytes"),
    # Pagination
    page: int = Query(1, ge=1, description="Page number"),
    page_size: int = Query(25, ge=1, le=250, description="Items per page"),
    # Sorting
    sort_by: str = Query("created_at", description="Field to sort by"),
    sort_order: str = Query("desc", description="Sort order (asc or desc)"),
    current_user: models.User = Depends(get_current_user),
) -> schemas.PaginatedFiles:
    """Get project files with advanced filtering, pagination, and sorting"""
    check_project_access(project_id, current_user, db)

    # Build base query with left join to get latest preprocessing task
    # Subquery to get the latest task created_at per file
    latest_task_date_subquery = (
        select(
            models.FilePreprocessingTask.file_id,
            func.max(models.FilePreprocessingTask.created_at).label("latest_task_date"),
        )
        .group_by(models.FilePreprocessingTask.file_id)
        .subquery()
    )

    # Subquery to get the latest task status per file
    # Use MIN(id) to ensure only one row per file_id when multiple tasks have the same timestamp
    latest_task_status_subquery = (
        select(
            models.FilePreprocessingTask.file_id,
            func.min(models.FilePreprocessingTask.id).label("latest_task_id"),
        )
        .join(
            latest_task_date_subquery,
            and_(
                models.FilePreprocessingTask.file_id
                == latest_task_date_subquery.c.file_id,
                models.FilePreprocessingTask.created_at
                == latest_task_date_subquery.c.latest_task_date,
            ),
        )
        .group_by(models.FilePreprocessingTask.file_id)
        .subquery()
    )

    # Get the status for the latest task (join with the task table using the min id)
    latest_task_status_with_status = (
        select(
            latest_task_status_subquery.c.file_id,
            models.FilePreprocessingTask.status.label("latest_status"),
        )
        .join(
            models.FilePreprocessingTask,
            models.FilePreprocessingTask.id
            == latest_task_status_subquery.c.latest_task_id,
        )
        .subquery()
    )

    query = (
        select(models.File)
        .where(models.File.project_id == project_id)
        .outerjoin(
            latest_task_status_with_status,
            models.File.id == latest_task_status_with_status.c.file_id,
        )
    )

    # Apply filters
    if search:
        query = query.where(models.File.file_name.ilike(f"%{search}%"))
    if file_type:
        query = query.where(models.File.file_type == file_type)
    if status:
        # Filter by preprocessing status based on latest task
        query = _apply_file_status_filter(query, status, latest_task_status_with_status)

    if file_creator is not None:
        query = query.where(models.File.file_creator == file_creator)
    if date_from:
        query = query.where(models.File.created_at >= date_from)
    if date_to:
        query = query.where(models.File.created_at <= date_to)
    if min_size is not None:
        query = query.where(models.File.file_size >= min_size)
    if max_size is not None:
        query = query.where(models.File.file_size <= max_size)

    # Get total count before pagination
    # Build the count query with the same joins and filters as the main query
    count_query = select(func.count(models.File.id)).where(
        models.File.project_id == project_id
    )

    # Apply the same filters to the count query
    if search:
        count_query = count_query.where(models.File.file_name.ilike(f"%{search}%"))
    if file_type:
        count_query = count_query.where(models.File.file_type == file_type)
    if status:
        # For status filter, we need to join with the status subquery
        count_query = count_query.outerjoin(
            latest_task_status_with_status,
            models.File.id == latest_task_status_with_status.c.file_id,
        )
        count_query = _apply_file_status_filter(
            count_query, status, latest_task_status_with_status
        )
    if file_creator is not None:
        count_query = count_query.where(models.File.file_creator == file_creator)
    if date_from:
        count_query = count_query.where(models.File.created_at >= date_from)
    if date_to:
        count_query = count_query.where(models.File.created_at <= date_to)
    if min_size is not None:
        count_query = count_query.where(models.File.file_size >= min_size)
    if max_size is not None:
        count_query = count_query.where(models.File.file_size <= max_size)

    total = db.execute(count_query).scalar() or 0

    # Apply sorting
    valid_sort_fields = {
        "file_name": models.File.file_name,
        "file_type": models.File.file_type,
        "file_size": models.File.file_size,
        "created_at": models.File.created_at,
    }
    sort_field = valid_sort_fields.get(sort_by, models.File.created_at)
    if sort_order.lower() == "asc":
        query = query.order_by(sort_field.asc())
    else:
        query = query.order_by(sort_field.desc())

    # Apply pagination
    offset = (page - 1) * page_size
    query = query.offset(offset).limit(page_size)

    # Eager load relationships to avoid N+1 queries in UI
    # Critical for performance with 50k+ files
    query = query.options(
        selectinload(models.File.preprocessing_tasks),
        selectinload(models.File.file_preprocessing_tasks),
        # Only load minimal fields for document relationships (avoid loading
        # full text). All four are loaded because the linked-resource check in
        # delete_file / file serialization touches all of them — a missing one
        # triggers a lazy N+1 per file.
        selectinload(models.File.documents_as_original).load_only(
            models.Document.id,
            models.Document.is_latest,
            models.Document.document_name,
        ),
        selectinload(models.File.documents_as_preprocessed).load_only(
            models.Document.id,
            models.Document.is_latest,
            models.Document.document_name,
        ),
    )

    files = list(db.execute(query).scalars().all())

    # Calculate total pages
    total_pages = (total + page_size - 1) // page_size if total > 0 else 0

    return schemas.PaginatedFiles(
        items=[schemas.File.model_validate(file) for file in files],
        total=total,
        page=page,
        page_size=page_size,
        total_pages=total_pages,
    )


@router.get("/stats", response_model=dict)
def get_file_stats(
    *,
    db: Session = Depends(get_db),
    project_id: int,
    file_creator: FileCreator | None = Query(None),
    current_user: models.User = Depends(get_current_user),
) -> dict:
    """Get file statistics for the project"""
    check_project_access(project_id, current_user, db)

    # Base query (shared by the aggregate and per-type stats below).
    # NB: previously the aggregate rebuilt its own WHERE with a buggy
    # `file_creator == file_creator if file_creator else True` that parsed as
    # `file_creator == (file_creator or True)`, returning 0 totals when no
    # file_creator filter was supplied. Reusing base_query fixes that.
    base_query = select(models.File).where(models.File.project_id == project_id)
    if file_creator is not None:
        base_query = base_query.where(models.File.file_creator == file_creator)

    # Get total count and size
    stats = db.execute(
        select(
            func.count(models.File.id).label("total_files"),
            func.sum(models.File.file_size).label("total_size"),
            func.count(distinct(models.File.file_hash)).label("unique_files"),
        ).select_from(base_query.subquery())
    ).first()

    # Get files by type
    type_query = select(
        models.File.file_type,
        func.count(models.File.id).label("count"),
        func.sum(models.File.file_size).label("size"),
    ).where(models.File.project_id == project_id)

    if file_creator is not None:
        type_query = type_query.where(models.File.file_creator == file_creator)

    type_stats = db.execute(type_query.group_by(models.File.file_type)).all()

    # Get recent files (last 7 days)
    week_ago = datetime.datetime.now(datetime.UTC) - datetime.timedelta(days=7)
    recent_query = select(func.count(models.File.id)).where(
        and_(models.File.project_id == project_id, models.File.created_at >= week_ago)
    )

    if file_creator is not None:
        recent_query = recent_query.where(models.File.file_creator == file_creator)

    recent_count = db.execute(recent_query).scalar()

    return {
        "total_files": (stats.total_files if stats else 0) or 0,
        "total_size": (stats.total_size if stats else 0) or 0,
        "unique_files": (stats.unique_files if stats else 0) or 0,
        "recent_files": recent_count or 0,
        "duplicates": ((stats.total_files if stats else 0) or 0)
        - ((stats.unique_files if stats else 0) or 0),
        "by_type": [
            {"type": t.file_type, "count": t.count, "size": t.size or 0}
            for t in type_stats
        ],
    }


@router.get("/{file_id}", response_model=schemas.File)
def get_project_file(
    *,
    db: Session = Depends(get_db),
    project_id: int,
    file_id: int,
    current_user: models.User = Depends(get_current_user),
) -> schemas.File:
    project: models.Project | None = db.execute(
        select(models.Project).where(models.Project.id == project_id)
    ).scalar_one_or_none()

    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    if not can_access_project(current_user, project):
        raise HTTPException(
            status_code=403, detail="Not authorized to access this project's files"
        )

    file: models.File | None = db.execute(
        select(models.File).where(
            models.File.project_id == project_id, models.File.id == file_id
        )
    ).scalar_one_or_none()

    if not file:
        raise HTTPException(status_code=404, detail="File not found")

    return schemas.File.model_validate(file)


@router.get("/{file_id}/content", response_class=StreamingResponse)
def get_project_file_content(
    *,
    db: Session = Depends(get_db),
    project_id: int,
    file_id: int,
    preview: bool = Query(False),
    current_user: models.User = Depends(get_current_user),
) -> StreamingResponse:
    """Retrieve the content of a file associated with a project."""
    project: models.Project | None = db.execute(
        select(models.Project).where(models.Project.id == project_id)
    ).scalar_one_or_none()

    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    if not can_access_project(current_user, project):
        raise HTTPException(
            status_code=403, detail="Not authorized to access this project's files"
        )

    file: models.File | None = db.execute(
        select(models.File).where(
            models.File.project_id == project_id, models.File.id == file_id
        )
    ).scalar_one_or_none()

    if not file:
        raise HTTPException(status_code=404, detail="File not found")

    # Stream the file from storage so a large PDF/image never has to be held in
    # memory in full before being sent to the client.
    #
    # Only a safe allowlist of types is served inline (browser-rendered);
    # everything else is forced as an attachment so user-uploaded HTML/SVG/JS
    # can't execute in the API origin (stored XSS — the JWT lives in
    # localStorage). SVG is deliberately excluded: it can carry <script>.
    safe_inline_types = {
        "application/pdf",
        "image/jpeg",
        "image/png",
        "image/gif",
        "image/webp",
    }
    serve_inline = preview and (file.file_type in safe_inline_types)
    disposition = "inline" if serve_inline else "attachment"
    # Audit every access to source PHI bytes. An attachment download (bytes
    # leaving the browser) is a FILE_DOWNLOAD; an inline browser-rendered
    # preview of the source document is a DOCUMENT_DOWNLOAD. The distinct
    # actions keep "downloaded" and "viewed inline" separable while ensuring
    # neither path is a blind spot.
    record_audit(
        AuditAction.DOCUMENT_DOWNLOAD if serve_inline else AuditAction.FILE_DOWNLOAD,
        actor=current_user,
        resource_type="file",
        resource_id=file.id,
        project_id=project_id,
        detail={"file_name": file.file_name, "disposition": disposition},
    )
    headers = {"Content-Disposition": f'{disposition}; filename="{file.file_name}"'}
    return StreamingResponse(
        stream_file(file.file_uuid),
        media_type=file.file_type,
        headers=headers,
    )


@router.post("", response_model=schemas.File)
def upload_file(
    *,
    db: Session = Depends(get_db),
    project_id: int,
    file: UploadFile = File(...),
    file_info: str = Form(...),
    current_user: models.User = Depends(get_current_user),
) -> schemas.File:
    """Upload a file with duplicate detection and *server-side MIME normalization*"""
    try:
        file_create = schemas.FileCreate.model_validate_json(file_info)
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Invalid JSON in file_info")
    except ValidationError as e:
        raise HTTPException(status_code=400, detail=str(e))

    check_project_access(project_id, current_user, db, permission="write")

    # Stream the upload ONCE to compute its hash + size and capture a small head
    # buffer for MIME sniffing, without ever loading the whole file into memory
    # (a 500MB upload would otherwise cost 500MB of RSS). Enforces the size cap
    # (413) during the pass. The file is rewound afterwards so it can be streamed
    # to storage below only if it's not a duplicate.
    file_hash, file_size, head = hash_measure_and_head(file)

    # --- Normalize MIME based on content + filename (fixes CSV mislabeled as vnd.ms-excel) ---
    # Prefer the originally submitted file name if present in the upload field; otherwise use the JSON's file_name.
    # `head` (first 8KB) is sufficient — detect_structured_mime only reads the
    # first 4096 bytes for magic-byte signatures.
    incoming_name = file.filename or file_create.file_name
    normalized_mime = detect_structured_mime(
        file_name=incoming_name,
        content=head,
        provided_mime=file.content_type or getattr(file_info, "file_type", None),
    )

    # Check duplicates by hash
    existing_file = db.execute(
        select(models.File).where(
            and_(
                models.File.project_id == project_id, models.File.file_hash == file_hash
            )
        )
    ).scalar_one_or_none()

    if existing_file:
        raise HTTPException(
            status_code=409,
            detail={
                "message": "File already exists",
                "existing_file": {
                    "id": existing_file.id,
                    "file_name": existing_file.file_name,
                    "created_at": existing_file.created_at.isoformat(),
                },
            },
        )

    # Finalize file info:
    # - Name: prefer actual upload name
    # - Type: normalized
    # - Store file_metadata from JSON if present
    file_create.file_name = incoming_name
    file_create.file_type = normalized_mime or "application/octet-stream"

    # Stream the (rewound) upload to storage — never buffering the whole file.
    file_uuid = save_upload_stream(file)

    new_file = models.File(
        **file_create.model_dump(
            exclude={"file_uuid", "file_size", "file_hash", "file_metadata"}
        ),
        project_id=project_id,
        file_uuid=file_uuid,
        file_size=file_size,
        file_hash=file_hash,
        file_metadata=getattr(file_create, "file_metadata", None),
    )

    db.add(new_file)
    try:
        db.commit()
    except Exception:
        # The bytes are already on disk/S3 with no DB row referencing them.
        # Roll back the session and remove the orphaned file so a failed
        # commit (constraint violation, DB error) doesn't leak storage.
        db.rollback()
        try:
            remove_file(file_uuid)
        except Exception:
            pass
        raise
    db.refresh(new_file)
    return schemas.File.model_validate(new_file)


def _load_full_table(file: models.File, file_content: bytes, metadata: dict):
    """Read a CSV/XLSX file into a full pandas DataFrame using import config.

    Mirrors the read logic in ``PreprocessingPipeline._process_table_file`` so
    that validation results match what preprocessing will actually see.
    """
    import pandas as pd

    has_header = metadata.get("has_header", True)
    header = 0 if has_header else None

    is_csv = file.file_type == models.FileType.TEXT_CSV or (
        file.file_name or ""
    ).lower().endswith(".csv")

    if is_csv:
        delimiter = metadata.get("delimiter") or ","
        encoding = metadata.get("encoding") or "utf-8"
        try:
            return pd.read_csv(
                io.BytesIO(file_content),
                encoding=encoding,
                delimiter=delimiter,
                header=header,
            )
        except Exception:
            # Fall back to lenient decoding, matching the pipeline behaviour.
            content = file_content.decode(encoding, errors="replace").encode("utf-8")
            return pd.read_csv(
                io.BytesIO(content),
                encoding="utf-8",
                delimiter=delimiter,
                header=header,
            )

    sheet = metadata.get("sheet")
    return pd.read_excel(
        io.BytesIO(file_content),
        sheet_name=sheet if sheet else 0,
        header=header,
    )


def _validate_id_column(file: models.File, metadata: dict) -> dict:
    """Check whether the configured case-ID column holds unique, non-empty values.

    Returns a structured result describing any duplicates so the client can tell
    the user exactly which IDs collide (and block saving) before preprocessing.
    """
    import pandas as pd

    case_id_column = metadata.get("case_id_column")

    # Nothing to validate: full-document imports or "(row number)" IDs are
    # always unique.
    if not case_id_column:
        return {"is_valid": True, "column_exists": True, "duplicates": []}

    file_content = get_file(file.file_uuid)
    try:
        df = _load_full_table(file, file_content, metadata)
    except Exception as exc:
        # Raw parse/storage errors can carry library internals or paths —
        # store the full exception in the error log and return only a safe
        # message with the error id.
        raise HTTPException(
            status_code=400,
            detail=internal_error_message(
                exc, prefix="Could not read file for validation"
            ),
        )

    if case_id_column not in df.columns:
        return {
            "is_valid": False,
            "column_exists": False,
            "total_rows": int(len(df)),
            "duplicates": [],
            "case_id_column": case_id_column,
        }

    col = df[case_id_column]
    counts = col.value_counts(dropna=False)
    duplicated = counts[counts > 1]

    duplicates = []
    for value, count in duplicated.items():
        is_empty = bool(pd.isna(value)) or (
            isinstance(value, str) and value.strip() == ""
        )
        duplicates.append(
            {
                "value": "" if pd.isna(value) else str(value),
                "count": int(count),
                "is_empty": is_empty,
            }
        )

    # Rows that share a non-unique ID (for a concise summary in the UI).
    duplicate_rows = int(duplicated.sum())

    return {
        "is_valid": len(duplicates) == 0,
        "column_exists": True,
        "case_id_column": case_id_column,
        "total_rows": int(len(df)),
        "duplicate_rows": duplicate_rows,
        # Cap the payload; the UI shows a "+N more" hint when truncated.
        "duplicates": duplicates[:50],
        "duplicate_value_count": len(duplicates),
    }


@router.post("/{file_id}/validate-id-column", response_model=dict)
def validate_id_column(
    *,
    db: Session = Depends(get_db),
    project_id: int,
    file_id: int,
    config: dict = Body(...),
    current_user: models.User = Depends(get_current_user),
) -> dict:
    """Validate that the chosen case-ID column is unique across the whole file.

    Called by the import-config modal before saving so duplicate IDs are caught
    up-front (with the offending values) instead of failing at preprocessing.
    """
    check_project_access(project_id, current_user, db)

    file = db.execute(
        select(models.File).where(
            models.File.project_id == project_id, models.File.id == file_id
        )
    ).scalar_one_or_none()
    if not file:
        raise HTTPException(status_code=404, detail="File not found")

    return _validate_id_column(file, config)


@router.post("/{file_id}/configure", response_model=schemas.File)
def configure_file_import(
    *,
    db: Session = Depends(get_db),
    project_id: int,
    file_id: int,
    config: dict = Body(...),
    current_user: models.User = Depends(get_current_user),
) -> schemas.File:
    """
    Set/update import config and preprocessing strategy for a file.
    """
    check_project_access(project_id, current_user, db, permission="write")

    file = db.execute(
        select(models.File).where(
            models.File.project_id == project_id, models.File.id == file_id
        )
    ).scalar_one_or_none()

    if not file:
        raise HTTPException(status_code=404, detail="File not found")

    # Accept "preprocessing_strategy" and/or "file_metadata"
    if "preprocessing_strategy" in config:
        file.preprocessing_strategy = config["preprocessing_strategy"]
    if "file_metadata" in config:
        file.file_metadata = config["file_metadata"]
    elif any(
        k in config
        for k in (
            "delimiter",
            "has_header",
            "row_split",
            "case_id_column",
            "text_columns",
        )
    ):
        file.file_metadata = {**(file.file_metadata or {}), **config}

    # Reject configs whose case-ID column is not unique across the whole file,
    # so the failure surfaces here instead of mid-preprocessing.
    if file.preprocessing_strategy == models.PreprocessingStrategy.ROW_BY_ROW and (
        file.file_metadata or {}
    ).get("case_id_column"):
        result = _validate_id_column(file, file.file_metadata or {})
        if not result["is_valid"]:
            db.rollback()
            raise HTTPException(status_code=422, detail=result)

    db.add(file)
    db.commit()
    db.refresh(file)

    return schemas.File.model_validate(file)


@router.get("/{file_id}/preview-rows", response_model=dict)
def preview_structured_file(
    *,
    db: Session = Depends(get_db),
    project_id: int,
    file_id: int,
    delimiter: str = Query(None, description="Delimiter for CSV"),
    encoding: str = Query("utf-8"),
    has_header: bool = Query(True),
    sheet: str = Query(None),
    max_rows: int = Query(10, ge=1, le=50),
    current_user: models.User = Depends(get_current_user),
) -> dict:
    """
    Return first N rows from CSV/XLSX for configuration/preview.
    Robust handling for:
      - Ambiguous MIME types (e.g., CSV uploaded as application/vnd.ms-excel)
      - Empty sheets
      - Blank headers / None cells
      - Invalid sheet names
      - Truncated CSV samples mid-line
      - Very long cells (clipped in preview)
      - Legacy XLS explicitly unsupported for preview
    """
    check_project_access(project_id, current_user, db)

    file = db.execute(
        select(models.File).where(
            models.File.project_id == project_id, models.File.id == file_id
        )
    ).scalar_one_or_none()
    if not file:
        raise HTTPException(status_code=404, detail="File not found")

    file_content = get_file(file.file_uuid)
    filename = (file.file_name or "").lower()

    # IMPORTANT: get Enum value string, not str(Enum)
    try:
        mime = (file.file_type.value or "").lower()
    except AttributeError:
        mime = (str(file.file_type or "")).lower()

    import csv as _csv
    import io as _io

    # ---- Helpers ----
    def _dedupe_headers(headers: list[str]) -> list[str]:
        seen: dict[str, int] = {}
        out: list[str] = []
        for h in headers:
            if h in seen:
                seen[h] += 1
                out.append(f"{h} ({seen[h]})")
            else:
                seen[h] = 1
                out.append(h)
        return out

    def _coerce_row(row: list) -> list:
        return [("" if v is None else v) for v in row]

    def _normalize_headers(
        raw_headers: list, width_fallback: int | None = None
    ) -> list[str]:
        if not raw_headers and width_fallback:
            headers = [f"Column {i + 1}" for i in range(width_fallback)]
        else:
            headers = []
            for idx, h in enumerate(raw_headers):
                if isinstance(h, str) and h.strip():
                    headers.append(h)
                else:
                    headers.append(f"Column {idx + 1}")
        return _dedupe_headers(headers)

    def _is_zip_xlsx(buf: bytes) -> bool:
        # XLSX is a ZIP: PK\x03\x04
        return len(buf) >= 4 and buf[:4] == b"PK\x03\x04"

    def _is_ole_xls(buf: bytes) -> bool:
        # Legacy XLS is OLE2/CFB: D0 CF 11 E0 A1 B1 1A 1E
        sig = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\x1e"
        return len(buf) >= len(sig) and buf[: len(sig)] == sig

    def _decide_format() -> str:
        """
        Decide 'csv' | 'xlsx' | 'xls' based on magic bytes and filename.
        Treat application/vnd.ms-excel as ambiguous (often CSV).
        """
        head = file_content[:16]

        # Strong magic checks first
        if _is_zip_xlsx(head) or filename.endswith(".xlsx"):
            return "xlsx"
        if _is_ole_xls(head) or filename.endswith(".xls"):
            return "xls"

        # CSV by extension or by elimination
        if filename.endswith(".csv"):
            return "csv"

        # MIME heuristics (browsers lie on Windows)
        if mime == "text/csv":
            return "csv"
        if mime == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
            # Double-check it's a real ZIP, else fall back to CSV
            return "xlsx" if _is_zip_xlsx(head) else "csv"
        if mime == "application/vnd.ms-excel":
            # Ambiguous; prefer magic: ZIP -> xlsx, OLE -> xls, else CSV
            if _is_zip_xlsx(head):
                return "xlsx"
            if _is_ole_xls(head):
                return "xls"
            return "csv"

        # Default: treat as CSV
        return "csv"

    # Limit single cell payload size
    CLIP = 5000
    decided = _decide_format()

    # ---------------- CSV ----------------
    def _parse_csv() -> dict:
        sample_bytes = file_content[:131072]  # 128 KiB
        try:
            sample = sample_bytes.decode(encoding, errors="replace")
            detected_encoding = encoding
        except Exception:
            sample = sample_bytes.decode("utf-8", errors="replace")
            detected_encoding = "utf-8"

        last_nl = sample.rfind("\n")
        if last_nl != -1:
            sample = sample[: last_nl + 1]

        # Detect delimiter if not provided
        sniffer = _csv.Sniffer()
        detected_delimiter = delimiter
        if not delimiter:
            try:
                dialect = sniffer.sniff(sample)
                detected_delimiter = dialect.delimiter
            except Exception:
                detected_delimiter = ","

        reader = _csv.reader(_io.StringIO(sample), delimiter=detected_delimiter)
        all_rows = list(reader)

        if not all_rows:
            return {
                "headers": [],
                "rows": [],
                "detected_delimiter": detected_delimiter,
                "detected_encoding": detected_encoding,
                "total_rows": 0,
                "truncated": False,
            }

        if has_header:
            raw_headers = _coerce_row(all_rows[0])
            headers = _normalize_headers(raw_headers)
            data_rows = all_rows[1 : 1 + max_rows]
        else:
            width = len(all_rows[0])
            headers = _normalize_headers([], width_fallback=width)
            data_rows = all_rows[:max_rows]

        preview_rows = []
        for r in data_rows:
            r = _coerce_row(r)
            preview_rows.append(
                [
                    (
                        c
                        if not isinstance(c, str) or len(c) <= CLIP
                        else (c[:CLIP] + "…")
                    )
                    for c in r
                ]
            )

        total_rows = len(all_rows) - (1 if has_header else 0)
        truncated = len(all_rows) > (max_rows + (1 if has_header else 0))

        return {
            "headers": headers,
            "rows": preview_rows,
            "detected_delimiter": detected_delimiter,
            "detected_encoding": detected_encoding,
            "total_rows": max(0, total_rows),
            "truncated": truncated,
        }

    if decided == "csv":
        return _parse_csv()

    # --------------- XLSX ----------------
    if decided == "xlsx":
        from zipfile import BadZipFile

        import openpyxl

        try:
            wb = openpyxl.load_workbook(
                _io.BytesIO(file_content), read_only=True, data_only=True
            )
        except (BadZipFile, KeyError, OSError, ValueError) as exc:
            # File was decided XLSX (usually by .xlsx extension) but the bytes
            # are not a valid OOXML/ZIP archive. Excel opens several such formats
            # that openpyxl cannot read, so disambiguate by magic bytes.
            head = file_content[:16]
            if _is_zip_xlsx(head):
                # Genuinely a ZIP but unreadable as a workbook. Ambiguous — a
                # benign corrupt upload, or a symptom of storage returning the
                # wrong/truncated bytes — so record it to the error log for admin
                # visibility and hand the correlation id to the user, while still
                # returning a clean 400 (not a 500).
                error_id = record_internal_error(exc, actor=current_user)
                raise HTTPException(
                    status_code=400,
                    detail=(
                        "This file is labeled as .xlsx but could not be read as "
                        "an Excel workbook. It may be corrupted. Please re-export "
                        "the file as .xlsx or .csv and try again. Quote this ID to "
                        f"your administrator for support: {error_id}"
                    ),
                ) from exc
            if _is_ole_xls(head):
                # Legacy binary .xls (OLE2) with an .xlsx name — Excel opens it,
                # openpyxl cannot.
                raise HTTPException(
                    status_code=400,
                    detail=(
                        "This looks like a legacy binary Excel file (.xls) even "
                        "though it is named .xlsx. Preview of binary .xls is not "
                        "supported — please re-save it as .xlsx or .csv."
                    ),
                )
            # Otherwise the content is textual (e.g. a CSV or HTML table renamed
            # to .xlsx) — fall back to CSV parsing so the preview still works.
            return _parse_csv()

        # Guard sheet selection
        if not sheet or sheet not in wb.sheetnames:
            sheet = wb.sheetnames[0]
        ws = wb[sheet]
        sheets = wb.sheetnames

        rows: list[list] = []
        take = max_rows + (1 if has_header else 0)
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            rows.append(_coerce_row(list(row)))
            if i + 1 >= take:
                break

        try:
            total_rows_raw = int(ws.max_row or 0)
        except Exception:
            total_rows_raw = 0
        total_rows = max(0, total_rows_raw - (1 if has_header else 0))
        truncated = total_rows > max_rows

        if not rows:
            return {
                "headers": [],
                "rows": [],
                "sheets": sheets,
                "total_rows": 0,
                "truncated": False,
            }

        if has_header:
            headers = _normalize_headers(rows[0])
            data_rows = rows[1:]
        else:
            width = len(rows[0])
            headers = _normalize_headers([], width_fallback=width)
            data_rows = rows

        preview_rows = []
        for r in data_rows[:max_rows]:
            preview_rows.append(
                [
                    (
                        c
                        if not isinstance(c, str) or len(c) <= CLIP
                        else (c[:CLIP] + "…")
                    )
                    for c in r
                ]
            )

        return {
            "headers": headers,
            "rows": preview_rows,
            "sheets": sheets,
            "total_rows": total_rows,
            "truncated": truncated,
        }

    # --------------- Legacy XLS (binary) ----------------
    if decided == "xls":
        raise HTTPException(
            status_code=400,
            detail=(
                "Preview for legacy .xls (binary) files is not supported. "
                "Please convert the file to .xlsx or .csv and try again."
            ),
        )

    raise HTTPException(
        status_code=400, detail="Preview not supported for this file type"
    )


@router.post("/check-duplicates", response_model=list[dict])
def check_duplicates(
    *,
    db: Session = Depends(get_db),
    project_id: int,
    files: list[dict] = Body(..., description="List of {filename, hash} objects"),
    current_user: models.User = Depends(get_current_user),
) -> list[dict]:
    """Check for duplicate files before uploading"""
    check_project_access(project_id, current_user, db)

    # Batch-load all matching files in one query (previously one SELECT per
    # file in the request body — N+1).
    hashes = [f["hash"] for f in files if f.get("hash")]
    existing_by_hash: dict[str, models.File] = {}
    if hashes:
        existing_by_hash = {
            f.file_hash: f
            for f in db.execute(
                select(models.File).where(
                    models.File.project_id == project_id,
                    models.File.file_hash.in_(hashes),
                )
            ).scalars()
        }

    results = []
    for file_info in files:
        existing = existing_by_hash.get(file_info["hash"])
        results.append(
            {
                "filename": file_info["filename"],
                "hash": file_info["hash"],
                "exists": existing is not None,
                "existing_file": {
                    "id": existing.id,
                    "file_name": existing.file_name,
                    "created_at": existing.created_at.isoformat(),
                }
                if existing
                else None,
            }
        )

    return results


@router.post("/dependencies", response_model=schemas.FileDependencies)
def get_file_dependencies(
    *,
    db: Session = Depends(get_db),
    project_id: int,
    payload: schemas.FileDependencyRequest,
    current_user: models.User = Depends(get_current_user),
) -> schemas.FileDependencies:
    """Summarize what a cascade delete of the given files would also remove.

    Resolves each file to the documents produced from it, then reports the same
    downstream impact (trials, groups, extraction results, evaluation metrics)
    as the document preview, plus the document count.
    """
    check_project_access(project_id, current_user, db, permission="write")
    if not payload.file_ids:
        return schemas.FileDependencies()

    doc_ids = list(
        db.execute(
            select(models.Document.id).where(
                models.Document.project_id == project_id,
                or_(
                    models.Document.original_file_id.in_(payload.file_ids),
                    models.Document.preprocessed_file_id.in_(payload.file_ids),
                ),
            )
        )
        .scalars()
        .all()
    )
    summary = compute_document_dependencies(db, project_id, doc_ids)
    return schemas.FileDependencies(documents=len(doc_ids), **summary)


@router.delete("/{file_id}", response_model=schemas.File)
def delete_file(
    *,
    db: Session = Depends(get_db),
    project_id: int,
    file_id: int,
    force: bool = Query(False, description="Force delete even if linked"),
    current_user: models.User = Depends(get_current_user),
) -> schemas.File:
    """Delete a file with safety checks"""
    check_project_access(project_id, current_user, db, permission="write")

    file = db.execute(
        select(models.File)
        .options(
            selectinload(models.File.documents_as_original),
            selectinload(models.File.documents_as_preprocessed),
            selectinload(models.File.preprocessing_tasks),
            selectinload(models.File.file_preprocessing_tasks),
        )
        .where(models.File.project_id == project_id, models.File.id == file_id)
    ).scalar_one_or_none()

    if not file:
        raise HTTPException(status_code=404, detail="File not found")

    doc_count = len(file.documents_as_original) + len(file.documents_as_preprocessed)

    # A file that's mid-preprocessing must not be deleted — the running task
    # would fail and leave orphaned rows. Block that regardless of `force`.
    # TERMINAL file tasks (failed/completed/cancelled) are just history and are
    # cleaned up below, so a *failed* preprocess no longer makes the file
    # undeletable (previously any file_preprocessing_task row blocked deletion).
    active_file_tasks = [
        ft
        for ft in file.file_preprocessing_tasks
        if ft.status
        in (
            models.PreprocessingStatus.PENDING,
            models.PreprocessingStatus.IN_PROGRESS,
        )
    ]
    if active_file_tasks:
        raise HTTPException(
            status_code=409,
            detail={
                "message": (
                    "This file is currently being preprocessed. Cancel the "
                    "preprocessing task before deleting the file."
                ),
                "links": {"active_preprocessing_tasks": len(active_file_tasks)},
            },
        )

    # Documents may be referenced by trials/evaluations, so still require force.
    if doc_count and not force:
        raise HTTPException(
            status_code=409,
            detail={
                "message": (
                    f"This file is linked to {doc_count} document(s). Delete those "
                    "first, or retry with force to remove them along with the file."
                ),
                "links": {"documents": doc_count},
            },
        )

    # With force, clear everything downstream of this file's documents — whole
    # trials, groups, evaluations, and residual result/metric rows — BEFORE the
    # document delete-orphan cascade runs below, so it doesn't trip the RESTRICT
    # FKs on trial results / evaluation metrics (which would otherwise 409).
    cascade_counts: dict[str, int] = {}
    orphaned_preprocessed_uuids: list[str] = []
    preprocessed_ids: set[int] = set()
    if force and doc_count:
        doc_ids = list(
            {d.id for d in file.documents_as_original}
            | {d.id for d in file.documents_as_preprocessed}
        )
        cascade_counts = cascade_clear_document_references(db, project_id, doc_ids)
        # The doomed documents point at system-generated preprocessed/OCR output
        # files — separate File rows (mostly S3 blobs). Collect them now, while
        # the relationships are loaded, so we can reclaim any that become orphaned
        # once the documents are deleted below (otherwise their rows AND bytes
        # would leak).
        preprocessed_ids = {
            d.preprocessed_file_id
            for d in (*file.documents_as_original, *file.documents_as_preprocessed)
            if d.preprocessed_file_id and d.preprocessed_file_id != file_id
        }

    # Remove the file's (terminal) preprocessing-task rows first:
    # file_preprocessing_tasks.file_id is a plain FK with no ON DELETE, so these
    # rows would otherwise block the file delete. Their documents cascade
    # (delete-orphan); a trial-referenced doc surfaces as the IntegrityError → 409
    # handler below. Then drop any parent PreprocessingTask left with no files.
    parent_task_ids = {ft.preprocessing_task_id for ft in file.file_preprocessing_tasks}
    for ft in list(file.file_preprocessing_tasks):
        db.delete(ft)
    if parent_task_ids:
        db.flush()
        for pid in parent_task_ids:
            remaining = db.scalar(
                select(func.count())
                .select_from(models.FilePreprocessingTask)
                .where(models.FilePreprocessingTask.preprocessing_task_id == pid)
            )
            if not remaining:
                parent = db.get(models.PreprocessingTask, pid)
                if parent:
                    db.delete(parent)

    # Reclaim now-orphaned preprocessed files (both the File row and its storage
    # blob). The documents that referenced them were just delete-orphaned via
    # their file tasks; flush so the reference check sees them gone, then delete
    # any preprocessed file no surviving document still points at.
    if preprocessed_ids:
        db.flush()
        for pf_id in preprocessed_ids:
            still_referenced = db.scalar(
                select(func.count())
                .select_from(models.Document)
                .where(
                    or_(
                        models.Document.original_file_id == pf_id,
                        models.Document.preprocessed_file_id == pf_id,
                    )
                )
            )
            if still_referenced:
                continue
            pf = db.get(models.File, pf_id)
            if pf is not None:
                if pf.file_uuid:
                    orphaned_preprocessed_uuids.append(pf.file_uuid)
                db.delete(pf)

    # Build the response while the instance is still attached, then delete the
    # DB row and commit BEFORE removing storage. If storage removal happened
    # first (the old order) a commit failure would orphan the DB row pointing
    # at deleted bytes; this order leaves at worst an orphaned file in storage
    # on a remove_file failure, which is recoverable.
    file_response = schemas.File.model_validate(file)
    file_uuid = file.file_uuid
    db.delete(file)
    try:
        db.commit()
    except IntegrityError:
        # Even with force=True the delete can hit a FK constraint — documents
        # reference the file via a NOT NULL, no-cascade FK, and trial results
        # reference documents via ON DELETE RESTRICT. Roll back and surface an
        # actionable 409 instead of letting it bubble up as a raw 500.
        db.rollback()
        raise HTTPException(
            status_code=409,
            detail={
                "message": (
                    "This file can't be deleted while documents or results still "
                    "reference it. Delete the dependent trials/documents first."
                ),
                "links": {
                    "documents": doc_count,
                },
            },
        )

    record_audit(
        AuditAction.DELETE,
        actor=current_user,
        resource_type="file",
        resource_id=file_id,
        project_id=project_id,
        detail={
            "forced": bool(force),
            "documents_deleted": doc_count,
            "preprocessed_files_deleted": len(orphaned_preprocessed_uuids),
            **cascade_counts,
        },
    )

    # Remove the storage blobs for the original file and every orphaned
    # preprocessed file, best-effort (a failure here only leaves recoverable
    # storage, never a dangling DB row).
    for uuid_to_remove in (file_uuid, *orphaned_preprocessed_uuids):
        if not uuid_to_remove:
            continue
        try:
            remove_file(uuid_to_remove)
        except FileNotFoundError:
            pass
        except Exception as e:
            logger.warning(
                "Failed to remove storage for file %s: %s", uuid_to_remove, e
            )

    return file_response


@router.post("/batch-delete", response_model=dict)
def batch_delete_files(
    *,
    db: Session = Depends(get_db),
    project_id: int,
    file_ids: list[int] = Body(...),
    force: bool = Body(False),
    current_user: models.User = Depends(get_current_user),
) -> dict:
    """Delete multiple files at once"""
    check_project_access(project_id, current_user, db, permission="write")

    # Cap the batch size to bound time/resource use (each delete does a
    # selectinload + storage removal). Matches the download-zip cap.
    max_files = 200
    if len(file_ids) > max_files:
        raise HTTPException(
            status_code=400,
            detail=f"Cannot delete more than {max_files} files at once "
            f"(requested {len(file_ids)}).",
        )

    deleted = []
    errors = []

    for file_id in file_ids:
        try:
            _ = delete_file(
                db=db,
                project_id=project_id,
                file_id=file_id,
                force=force,
                current_user=current_user,
            )
            deleted.append(file_id)
        except HTTPException as e:
            # Our own controlled message (e.g. "file is linked, use force").
            errors.append({"file_id": file_id, "error": e.detail})
        except Exception as e:
            # Unexpected failure (DB/storage) — surface a correlation id, not the
            # raw exception text, so internals don't leak to the client.
            error_id = record_internal_error(e, actor=current_user)
            errors.append(
                {
                    "file_id": file_id,
                    "error_id": error_id,
                    "error": "Failed to delete file (see error id).",
                }
            )

    return {
        "deleted": deleted,
        "errors": errors,
        "total_deleted": len(deleted),
        "total_errors": len(errors),
    }


@router.post("/download-zip")
def download_files_as_zip(
    *,
    db: Session = Depends(get_db),
    project_id: int,
    file_ids: list[int] = Body(...),
    include_metadata: bool = Body(True),
    current_user: models.User = Depends(get_current_user),
) -> Response:
    """Download multiple files as a ZIP archive.

    Plain (non-async) ``def`` so FastAPI runs it in a threadpool — the handler
    does blocking sync I/O (``get_file`` reads from disk/S3 and the DB session
    is synchronous), which would otherwise block the event loop for the entire
    request.
    """
    check_project_access(project_id, current_user, db)

    # Cap the number of files to bound memory/time for ZIP assembly.
    max_files = 200
    if len(file_ids) > max_files:
        raise HTTPException(
            status_code=400,
            detail=f"Cannot zip more than {max_files} files at once "
            f"(requested {len(file_ids)}).",
        )

    # Batch-load all requested files in one query instead of one per id (N+1).
    files = (
        db.execute(
            select(models.File).where(
                models.File.project_id == project_id,
                models.File.id.in_(file_ids),
            )
        )
        .scalars()
        .all()
    )
    # Preserve the requested order.
    files_by_id = {f.id: f for f in files}
    ordered_files = [files_by_id[i] for i in file_ids if i in files_by_id]

    # Create ZIP file in memory
    zip_buffer = io.BytesIO()

    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
        files_metadata = []

        for file in ordered_files:
            try:
                # Get file content
                file_content = get_file(file.file_uuid)

                # Add file to ZIP
                zip_file.writestr(file.file_name, file_content)

                # Collect metadata
                if include_metadata:
                    files_metadata.append(
                        {
                            "id": file.id,
                            "file_name": file.file_name,
                            "file_type": file.file_type,
                            "file_size": file.file_size,
                            "file_hash": file.file_hash,
                            "description": file.description,
                            "created_at": file.created_at.isoformat()
                            if file.created_at
                            else None,
                            "updated_at": file.updated_at.isoformat()
                            if file.updated_at
                            else None,
                        }
                    )

            except Exception as e:
                logger.error("Error adding file %s to ZIP: %s", file.file_name, e)
                continue

        # Add metadata file if requested
        if include_metadata and files_metadata:
            metadata_json = json.dumps(
                {
                    "project_id": project_id,
                    "export_date": datetime.datetime.now(datetime.UTC).isoformat(),
                    "total_files": len(files_metadata),
                    "files": files_metadata,
                },
                indent=2,
            )
            zip_file.writestr("metadata.json", metadata_json)

    # Prepare response
    zip_buffer.seek(0)
    filename = f"project_{project_id}_files_{datetime.datetime.now(datetime.UTC).strftime('%Y%m%d_%H%M%S')}.zip"

    record_audit(
        AuditAction.EXPORT,
        actor=current_user,
        resource_type="file",
        project_id=project_id,
        detail={"files": len(file_ids), "format": "zip"},
    )
    return Response(
        content=zip_buffer.getvalue(),
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/move")
def move_files(
    *,
    db: Session = Depends(get_db),
    project_id: int,
    file_ids: list[int] = Body(...),
    target_project_id: int = Body(...),
    current_user: models.User = Depends(get_current_user),
) -> dict:
    """Move files to another project.

    Plain (non-async) ``def`` so FastAPI runs it in a threadpool — the handler
    uses a synchronous DB session, which would otherwise block the event loop.
    """
    # Check access to both source and target projects
    check_project_access(project_id, current_user, db, permission="write")
    check_project_access(target_project_id, current_user, db, permission="write")

    if project_id == target_project_id:
        raise HTTPException(
            status_code=400, detail="Source and target projects cannot be the same"
        )

    # Cap the batch: the move loop issues a DB query per file (N+1). An
    # unbounded list would tie up a worker and saturate the connection pool.
    # Mirrors batch_delete_files / download_files_as_zip (cap = 200).
    max_files = 200
    if len(file_ids) > max_files:
        raise HTTPException(
            status_code=400,
            detail=f"Cannot move more than {max_files} files at once "
            f"(requested {len(file_ids)}).",
        )

    # Pre-flight: collect the IDs of all documents that would move, then block
    # the move if any are still referenced by trials or document sets in the
    # source project. Moving the documents to another project would otherwise
    # leave those references pointing across projects, breaking project
    # isolation (a trial in project A silently referencing documents in B).
    moving_doc_ids = [
        row[0]
        for row in db.execute(
            select(models.Document.id).where(
                or_(
                    models.Document.original_file_id.in_(file_ids),
                    models.Document.preprocessed_file_id.in_(file_ids),
                )
            )
        ).all()
    ]

    if moving_doc_ids:
        # Trials referencing these documents via the document_ids JSON column.
        source_trials = (
            db.execute(
                select(models.Trial.document_ids).where(
                    models.Trial.project_id == project_id
                )
            )
            .scalars()
            .all()
        )
        moving_doc_id_set = set(moving_doc_ids)
        referenced_by_trials = {
            doc_id
            for trial_doc_ids in source_trials
            if trial_doc_ids
            for doc_id in trial_doc_ids
            if doc_id in moving_doc_id_set
        }

        # Document sets referencing these documents via the association table.
        referenced_by_sets = set(
            row[0]
            for row in db.execute(
                select(document_set_association.c.document_id).where(
                    document_set_association.c.document_id.in_(moving_doc_ids)
                )
            ).all()
        )

        referenced = referenced_by_trials | referenced_by_sets
        if referenced:
            raise HTTPException(
                status_code=409,
                detail={
                    "message": (
                        "Cannot move files: some documents are referenced by "
                        "trials or document sets in this project. Remove the "
                        "references first."
                    ),
                    "referenced_document_ids": sorted(referenced),
                },
            )

    moved_count = 0
    errors = []

    # Cache of source-config-id → cloned-target-config so multiple files/docs
    # sharing a preprocessing config reuse one clone in the target project.
    config_clone_map: dict[int, int] = {}

    def _target_config_id(src_config_id: int) -> int:
        cached = config_clone_map.get(src_config_id)
        if cached is not None:
            return cached
        src = db.get(models.PreprocessingConfiguration, src_config_id)
        clone = models.PreprocessingConfiguration(
            project_id=target_project_id,
            name=src.name,
            description=src.description,
            additional_settings=dict(src.additional_settings or {}),
        )
        db.add(clone)
        db.flush()  # assign clone.id
        config_clone_map[src_config_id] = clone.id
        return clone.id

    for file_id in file_ids:
        try:
            file = db.execute(
                select(models.File).where(
                    models.File.id == file_id, models.File.project_id == project_id
                )
            ).scalar_one_or_none()

            if not file:
                errors.append({"file_id": file_id, "error": "File not found"})
                continue

            # Update the file's project
            file.project_id = target_project_id

            # Update any related documents
            documents = (
                db.execute(
                    select(models.Document).where(
                        or_(
                            models.Document.original_file_id == file_id,
                            models.Document.preprocessed_file_id == file_id,
                        )
                    )
                )
                .scalars()
                .all()
            )

            for doc in documents:
                doc.project_id = target_project_id
                # Sever cross-project lineage so that later deleting the SOURCE
                # project can't cascade-delete these now-target documents:
                #  - preprocessing_config_id is NOT NULL and its config lives in
                #    (and is delete-orphan'd with) the source project, so repoint
                #    it at a clone in the target project;
                #  - file_preprocessing_task_id's relationship is
                #    cascade="all, delete-orphan", so leaving it pointed at the
                #    source FilePreprocessingTask would let the source project's
                #    deletion wipe these moved documents. Null it (provenance to
                #    the original task is dropped; the file itself has moved).
                doc.preprocessing_config_id = _target_config_id(
                    doc.preprocessing_config_id
                )
                doc.file_preprocessing_task_id = None

            moved_count += 1

        except Exception as e:
            # Surface a correlation id rather than raw exception text.
            error_id = record_internal_error(e, actor=current_user)
            errors.append(
                {
                    "file_id": file_id,
                    "error_id": error_id,
                    "error": "Failed to move file (see error id).",
                }
            )

    db.commit()

    return {"moved": moved_count, "errors": errors, "total_requested": len(file_ids)}


@router.get("/check-links/{file_id}")
def check_file_links(
    *,
    db: Session = Depends(get_db),
    project_id: int,
    file_id: int,
    current_user: models.User = Depends(get_current_user),
) -> dict:
    """Check if a file has any linked resources"""
    check_project_access(project_id, current_user, db)

    file = db.execute(
        select(models.File)
        .options(
            selectinload(models.File.documents_as_original),
            selectinload(models.File.documents_as_preprocessed),
            selectinload(models.File.preprocessing_tasks),
            selectinload(models.File.file_preprocessing_tasks),
        )
        .where(models.File.project_id == project_id, models.File.id == file_id)
    ).scalar_one_or_none()

    if not file:
        raise HTTPException(status_code=404, detail="File not found")

    return {
        "file_id": file_id,
        "is_linked": bool(
            file.documents_as_original
            or file.documents_as_preprocessed
            or file.preprocessing_tasks
            or file.file_preprocessing_tasks
        ),
        "links": {
            "documents_as_original": len(file.documents_as_original),
            "documents_as_preprocessed": len(file.documents_as_preprocessed),
            "preprocessing_tasks": len(file.preprocessing_tasks),
            "file_preprocessing_tasks": len(file.file_preprocessing_tasks),
        },
    }
